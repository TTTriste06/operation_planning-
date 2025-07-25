import re
import pandas as pd
import streamlit as st
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from calendar import monthrange
from datetime import datetime


def extract_wafer_with_grossdie_raw(main_plan_df: pd.DataFrame, df_grossdie: pd.DataFrame) -> pd.DataFrame:
    """
    直接将晶圆品名与 df_grossdie 的“规格”列做匹配，如果匹配上则取该行的“GROSS DIE”。

    参数：
        df_grossdie: 原始 grossdie 表（不可清洗）
        main_plan_df: 主计划表，包含“晶圆品名”

    返回：
        DataFrame: 包含“晶圆品名”和“单片数量”的 DataFrame
    """
    # 提取唯一晶圆品名
    wafer_names = (
        main_plan_df["晶圆品名"]
        .dropna()
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .reset_index(drop=True)
    )
    df_result = pd.DataFrame({"晶圆品名": wafer_names})

    # 匹配逻辑：晶圆品名是否出现在 grossdie 的规格列中
    def match_grossdie(wafer_name):
        matched = df_grossdie[df_grossdie["规格"] == wafer_name]
        if not matched.empty:
            return matched.iloc[0]["GROSS DIE"]
        return None

    df_result["单片数量"] = df_result["晶圆品名"].apply(match_grossdie)

    return df_result


def append_inventory_columns(df_unique_wafer: pd.DataFrame, main_plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    将每个晶圆品名在 main_plan_df 中对应的 InvWaf 与 InvPart 求和后，填入 df_unique_wafer。

    参数：
        df_unique_wafer: 包含唯一“晶圆品名”的 DataFrame
        main_plan_df: 包含完整数据（包含“晶圆品名”, "InvWaf", "InvPart"）

    返回：
        更新后的 df_unique_wafer，新增列：InvWaf, InvPart
    """
    # 只保留必要列并转换类型
    wafer_inventory = (
        main_plan_df[["晶圆品名", "InvWaf", "InvPart"]]
        .copy()
        .dropna(subset=["晶圆品名"])
    )
    wafer_inventory["晶圆品名"] = wafer_inventory["晶圆品名"].astype(str).str.strip()

    # 求和：以晶圆品名为索引聚合
    inventory_sum = wafer_inventory.groupby("晶圆品名", as_index=False)[["InvWaf", "InvPart"]].sum()

    # 合并回 df_unique_wafer
    df_unique_wafer = df_unique_wafer.copy()
    df_unique_wafer["晶圆品名"] = df_unique_wafer["晶圆品名"].astype(str).str.strip()

    df_merged = pd.merge(df_unique_wafer, inventory_sum, on="晶圆品名", how="left")

    return df_merged


def append_wafer_inventory_by_warehouse(df_unique_wafer: pd.DataFrame, wafer_inventory_df: pd.DataFrame) -> pd.DataFrame:
    """
    根据“晶圆品名”匹配 wafer_inventory_df 中的“WAFER品名”，
    并将其数量按“仓库名称”展开成多列，汇总填入 df_unique_wafer。
    """
    # 标准化字段
    wafer_inventory_df = wafer_inventory_df.copy()
    wafer_inventory_df["WAFER品名"] = wafer_inventory_df["WAFER品名"].astype(str).str.strip()
    wafer_inventory_df["仓库名称"] = wafer_inventory_df["仓库名称"].astype(str).str.strip()

    # 过滤出匹配的晶圆品名
    matched_inventory = wafer_inventory_df[
        wafer_inventory_df["WAFER品名"].isin(df_unique_wafer["晶圆品名"])
    ].copy()

    # 将“数量”确保是数字
    matched_inventory["数量"] = pd.to_numeric(matched_inventory["数量"], errors="coerce").fillna(0)

    # 透视表：按“晶圆品名”和“仓库名称”聚合数量
    pivot_inventory = matched_inventory.pivot_table(
        index="WAFER品名",
        columns="仓库名称",
        values="数量",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # 重命名 WAFER品名 → 晶圆品名，方便 merge
    pivot_inventory = pivot_inventory.rename(columns={"WAFER品名": "晶圆品名"})

    # 合并到原表
    df_result = pd.merge(df_unique_wafer, pivot_inventory, on="晶圆品名", how="left")

    return df_result

def merge_wafer_inventory_columns(ws: Worksheet, df: pd.DataFrame):
    """
    查找所有以“仓”结尾的列，在第一行合并并写入“晶圆库存”。

    参数：
        ws: openpyxl 的 Worksheet 对象（例如“主计划”sheet）
        df: 对应 DataFrame，用于定位列位置
    """
    # 1. 找出所有以“仓”结尾的列名
    inventory_cols = [col for col in df.columns if str(col).strip().endswith("仓")]
    if not inventory_cols:
        return  # 没有匹配到“仓”列，无需处理

    # 2. 获取这些列在 DataFrame 中的索引位置（从0开始）转为 Excel 列号（从1开始）
    start_col_idx = df.columns.get_loc(inventory_cols[0]) + 1
    end_col_idx = df.columns.get_loc(inventory_cols[-1]) + 1

    # 3. 获取列字母（如 E, F）
    start_letter = get_column_letter(start_col_idx)
    end_letter = get_column_letter(end_col_idx)

    # 4. 合并单元格并写入标题“晶圆库存”
    title_cell = ws.cell(row=1, column=start_col_idx, value="晶圆库存")
    ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
    
    # 5. 样式设置
    title_cell.alignment = Alignment(horizontal="center", vertical="center")


def append_cp_wip_total(df_unique_wafer: pd.DataFrame, df_cp_wip: pd.DataFrame) -> pd.DataFrame:
    """
    将 CP 在制表中的“未交”总数按“晶圆型号”匹配到 df_unique_wafer 的“晶圆品名”列。

    参数：
        df_unique_wafer: 包含唯一“晶圆品名”的 DataFrame
        df_cp_wip: CP 在制表，必须包含“晶圆型号”和“未交”

    返回：
        带有“CP在制（Total）”列的新 DataFrame
    """
    # 清理字段
    df_cp_wip = df_cp_wip.copy()
    df_cp_wip["晶圆型号"] = df_cp_wip["晶圆型号"].astype(str).str.strip()
    df_cp_wip["未交"] = pd.to_numeric(df_cp_wip["未交"], errors="coerce").fillna(0)

    # 按“晶圆型号”汇总未交数量
    cp_total = df_cp_wip.groupby("晶圆型号", as_index=False)["未交"].sum()
    cp_total = cp_total.rename(columns={"晶圆型号": "晶圆品名", "未交": "CP在制（Total）"})

    # 合并回 df_unique_wafer
    df_result = pd.merge(df_unique_wafer, cp_total, on="晶圆品名", how="left")

    return df_result

def merge_cp_wip_column(ws: Worksheet, df: pd.DataFrame):
    """
    在 Excel 中对“CP在制（Total）”这一列合并上方单元格，写入“在制CP晶圆”标题。
    
    参数：
        ws: openpyxl 的工作表对象
        df: DataFrame（用于查找列位置）
    """
    # 确保列存在
    if "CP在制（Total）" not in df.columns:
        return

    # 获取该列在 DataFrame 中的列索引（从 0 开始），转为 Excel 列号（从 1 开始）
    col_idx = df.columns.get_loc("CP在制（Total）") + 1
    col_letter = get_column_letter(col_idx)

    # 合并第一行并写入标题
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx)
    cell = ws.cell(row=1, column=col_idx)
    cell.value = "在制CP晶圆"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_monthly_wo_from_weekly_fab(df_unique_wafer: pd.DataFrame, df_fab_summary: pd.DataFrame) -> pd.DataFrame:
    """
    将 df_fab_summary 中的周产出列按月份汇总为“yyyy-mm WO”列，并合并到 df_unique_wafer。
    """
    df = df_unique_wafer.copy()
    df_fab = df_fab_summary.copy()

    # 标准化晶圆品名列
    df["晶圆品名"] = df["晶圆品名"].astype(str).str.strip()
    df_fab["晶圆型号"] = df_fab["晶圆型号"].astype(str).str.strip()

    # 识别周列（排除“晶圆型号”, “FAB”等）
    known_cols = ["晶圆型号", "FAB"]
    week_cols = [col for col in df_fab.columns if col not in known_cols]

    # 提取“yyyy-mm” → { "2025-07": ["2025-07 WK1(1–7)", ...] }
    month_to_weeks = {}
    for col in week_cols:
        match = re.match(r"(\d{4}-\d{2})", col)
        if match:
            month = match.group(1)
            month_to_weeks.setdefault(month, []).append(col)

    # 计算每个月的总 WO 列
    monthly_agg = pd.DataFrame()
    monthly_agg["晶圆型号"] = df_fab["晶圆型号"]

    for month, cols in month_to_weeks.items():
        monthly_agg[f"{month} WO"] = df_fab[cols].sum(axis=1)

    # 重命名用于合并
    monthly_agg = monthly_agg.rename(columns={"晶圆型号": "晶圆品名"})

    # 合并到 df
    df_result = pd.merge(df, monthly_agg, on="晶圆品名", how="left")

    return df_result

def merge_monthly_fab_wo_columns(ws: Worksheet, df: pd.DataFrame):
    """
    对所有“xxxx-xx WO”结尾的列，在第一行合并并写“Fab预计晶圆产出数量”。

    参数：
        ws: openpyxl 的 worksheet
        df: DataFrame 用于定位列索引
    """
    # 识别所有“xxx WO”结尾的列
    wo_cols = [col for col in df.columns if str(col).strip().endswith(" WO")]
    if not wo_cols:
        return

    start_col_idx = df.columns.get_loc(wo_cols[0]) + 1
    end_col_idx = df.columns.get_loc(wo_cols[-1]) + 1

    # 合并
    ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
    cell = ws.cell(row=1, column=start_col_idx)
    cell.value = "Fab预计晶圆产出数量"

    # 样式设置
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_monthly_demand_from_fg_plan(df_unique_wafer: pd.DataFrame, main_plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    将 main_plan_df 中“YYYY-MM成品投单计划”字段按晶圆品名汇总为“YYYY-MM需求”，并差分添加至 df_unique_wafer：
        - 第一个月：成品投单计划 - InvPart
        - 后续月份：当前月 - 上月（可为负）
    """
    df = df_unique_wafer.copy()
    df["晶圆品名"] = df["晶圆品名"].astype(str).str.strip()
    main_plan_df["晶圆品名"] = main_plan_df["晶圆品名"].astype(str).str.strip()

    # ✅ 匹配所有 "YYYY-MM成品投单计划" 列
    pattern = re.compile(r"^(\d{4})-(\d{2})成品投单计划$")
    plan_cols = [col for col in main_plan_df.columns if pattern.match(col)]

    if not plan_cols:
        raise ValueError("❌ 未找到“YYYY-MM成品投单计划”列")

    # ✅ 按日期升序排序（保证顺序正确）
    sorted_plan_cols = sorted(
        plan_cols,
        key=lambda col: pd.to_datetime(pattern.match(col).group(1) + "-" + pattern.match(col).group(2))
    )

    # ✅ 聚合计划值
    grouped = main_plan_df[["晶圆品名"] + sorted_plan_cols].copy()
    grouped[sorted_plan_cols] = grouped[sorted_plan_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    grouped = grouped.groupby("晶圆品名", as_index=False)[sorted_plan_cols].sum()

    # ✅ 差分需求
    diff_df = grouped[["晶圆品名"]].copy()
    for i, col in enumerate(sorted_plan_cols):
        if i == 0:
            diff_df[col] = grouped[col]
        else:
            prev_col = sorted_plan_cols[i - 1]
            diff_df[col] = grouped[col] - grouped[prev_col]

    # ✅ 合并进 df_unique_wafer 并减去 InvPart
    temp_df = pd.merge(df, diff_df, on="晶圆品名", how="left")

    first_col = sorted_plan_cols[0]
    if "InvPart" not in temp_df.columns:
        raise ValueError("❌ df_unique_wafer 中缺少 InvPart 列")

    temp_df[first_col] = temp_df[first_col] - temp_df["InvPart"]

    # ✅ 重命名为 “YYYY-MM需求”
    rename_dict = {col: f"{col[:-6]}需求" for col in sorted_plan_cols}  # 去掉“成品投单计划”6字
    temp_df.rename(columns=rename_dict, inplace=True)

    # ✅ 保留三位小数
    for col in rename_dict.values():
        temp_df[col] = temp_df[col].round(3)

    return temp_df


def merge_fg_plan_columns(ws: Worksheet, df: pd.DataFrame):
    """
    将所有“yyyy-mm需求”列中来源于成品投单计划的部分合并在第1行，写入“成品投单计划”
    """
    pattern = re.compile(r"^20\d{2}-\d{2}需求$")
    demand_cols = [col for col in df.columns if pattern.match(str(col))]

    if not demand_cols:
        return

    # 默认将最后一段连续的“需求”列作为“成品投单计划”
    end_idx = df.columns.get_loc(demand_cols[-1])
    start_idx = end_idx
    for i in reversed(range(end_idx)):
        if pattern.match(str(df.columns[i])):
            start_idx = i
        else:
            break

    start_col = start_idx + 1
    end_col = end_idx + 1

    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col)
    cell.value = "成品投单计划"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def fill_columns_c_and_right_with_zero(df: pd.DataFrame) -> pd.DataFrame:
    """
    将 DataFrame 中第3列及其右侧所有空单元格填入0。
    """
    df_copy = df.copy()
    start_col = 2  # 第3列的索引（从0开始）
    df_copy.iloc[:, start_col:] = df_copy.iloc[:, start_col:].fillna(0)
    return df_copy


def merge_allocation_header(ws: Worksheet):
    """
    将所有“yyyy-mm分配”列的标题行合并，并写上“晶圆分配（颗）”
    """
    pattern = re.compile(r"^20\d{2}-\d{2}分配$")
    header_row = 2
    matched_cols = []

    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value and pattern.match(str(cell.value)):
            matched_cols.append(col_idx)

    if not matched_cols:
        return

    start_col = matched_cols[0]
    end_col = matched_cols[-1]
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)

    ws.merge_cells(f"{start_letter}1:{end_letter}1")
    cell = ws.cell(row=1, column=start_col)
    cell.value = "晶圆分配（颗）"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_monthly_gap_columns(df_unique_wafer: pd.DataFrame) -> pd.DataFrame:
    """
    在 df_unique_wafer 后添加每个月的 “yyyy-mm缺口” 列：
    缺口 = (需求 - 分配) / 单片数量，保留三位小数
    """
    df = df_unique_wafer.copy()

    # 匹配所有 “2025-xx需求” 列
    pattern = re.compile(r"^(20\d{2}-\d{2})需求$")
    demand_cols = [col for col in df.columns if pattern.match(str(col))]

    for demand_col in demand_cols:
        month = pattern.match(demand_col).group(1)
        allocation_col = f"{month}分配"
        gap_col = f"{month}缺口"

        if allocation_col in df.columns:
            single_die = df["单片数量"].replace({0: float("nan")})
            df[gap_col] = (df[demand_col] - df[allocation_col]) / single_die
            df[gap_col] = df[gap_col].fillna(0).round(3)
        else:
            raise ValueError(f"❌ 缺少对应的分配列：{allocation_col}")

    return df


def merge_monthly_gap_columns(ws: Worksheet):
    """
    将所有“yyyy-mm缺口”列合并为一个上层标题“晶圆缺口计算（片）”，位于第一行
    """
    pattern = re.compile(r"^20\d{2}-\d{2}缺口$")
    header_row = 2
    matched_cols = []

    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value and pattern.match(str(cell.value)):
            matched_cols.append(col_idx)

    if not matched_cols:
        return

    start_col = matched_cols[0]
    end_col = matched_cols[-1]
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)

    ws.merge_cells(f"{start_letter}1:{end_letter}1")
    cell = ws.cell(row=1, column=start_col)
    cell.value = "晶圆缺口计算（片）"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_fab_warehouse_quantity(df_unique_wafer: pd.DataFrame, sh_fabout_dict: dict) -> pd.DataFrame:
    """
    从 SH_fabout 中提取所有晶圆品名的 FABOUT_QTY 总和，合并入 df_unique_wafer 的 'Fab warehouse' 列。
    """
    from collections import defaultdict

    # 初始化总量累加器
    total_fabout = defaultdict(float)

    for sheet_name, df in sh_fabout_dict.items():
        if "CUST_PARTNAME" not in df.columns or "FABOUT_QTY" not in df.columns:
            print(f"❌ 表 {sheet_name} 缺少必要字段，跳过")
            continue

        # 标准化
        df = df.copy()
        df["CUST_PARTNAME"] = df["CUST_PARTNAME"].astype(str).str.strip()
        df["FABOUT_QTY"] = pd.to_numeric(df["FABOUT_QTY"], errors="coerce").fillna(0)

        grouped = df.groupby("CUST_PARTNAME")["FABOUT_QTY"].sum()

        for partname, qty in grouped.items():
            total_fabout[partname] += qty

    # 转换为 DataFrame
    fab_df = pd.DataFrame(list(total_fabout.items()), columns=["晶圆品名", "Fab warehouse"])
    fab_df["晶圆品名"] = fab_df["晶圆品名"].astype(str).str.strip()

    # 匹配目标列也做清洗
    df_unique_wafer = df_unique_wafer.copy()
    df_unique_wafer["晶圆品名"] = df_unique_wafer["晶圆品名"].astype(str).str.strip()

    # 合并
    df_result = pd.merge(df_unique_wafer, fab_df, on="晶圆品名", how="left")
    return df_result

def merge_fab_warehouse_column(ws: Worksheet, df: pd.DataFrame):
    """
    在 Excel 中对“Fab warehouse”列合并第一行并写入“Fabout”作为分组标题。

    参数：
        ws: openpyxl 工作表对象
        df: DataFrame，用于定位该列位置
    """
    if "Fab warehouse" not in df.columns:
        return  # 列不存在，跳过

    # 获取该列索引（Excel 从 1 开始）
    col_idx = df.columns.get_loc("Fab warehouse") + 1
    col_letter = get_column_letter(col_idx)

    # 合并单元格（仅 1 列）
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx)

    # 写入标题
    cell = ws.cell(row=1, column=col_idx)
    cell.value = "Fabout"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def allocate_fg_demand_monthly(df_unique_wafer: pd.DataFrame, start_date: datetime) -> pd.DataFrame:
    """
    根据“YYYY-MM需求”列，逐月分配“YYYY-MM分配”，
    并使用每月的 WO 和仓库存进行总量控制。

    参数:
        df_unique_wafer: 包含“需求”、“WO”、“仓库存”等字段的 DataFrame
        start_date: datetime 类型，定义第一个月

    返回:
        包含“YYYY-MM分配”列的 df
    """
    first_date = datetime(start_date.year, start_date.month, 1)
    df = df_unique_wafer.copy()

    # ✅ 获取所有“YYYY-MM需求”列
    pattern = re.compile(r"^(\d{4})-(\d{2})需求$")
    demand_cols = [col for col in df.columns if pattern.match(col)]
    if not demand_cols:
        raise ValueError("❌ 未找到任何“YYYY-MM需求”列")

    # ✅ 按时间升序排序
    demand_months = sorted(
        [(col, pd.to_datetime(f"{m.group(1)}-{m.group(2)}"))
         for col in demand_cols if (m := pattern.match(col))],
        key=lambda x: x[1]
    )
    sorted_demand_cols = [col for col, _ in demand_months]
    sorted_month_strs = [dt.strftime("%Y-%m") for _, dt in demand_months]
    allocation_cols = [f"{ym}分配" for ym in sorted_month_strs]

    # ✅ 初始化所有分配列
    for col in allocation_cols:
        df[col] = 0.0

    # ✅ 提取 WO 列和对应时间
    wo_pattern = re.compile(r"^(\d{4})-(\d{2}) WO$")
    wo_cols = [(col, pd.to_datetime(f"{m.group(1)}-{m.group(2)}"))
               for col in df.columns if (m := wo_pattern.match(col))]
    wo_cols.sort(key=lambda x: x[1])

    # ✅ 分配逻辑
    for idx, row in df.iterrows():
        rest_prev = 0
        wafer_unit = pd.to_numeric(row.get("单片数量", 1.0), errors="coerce") or 1.0

        for i, ym in enumerate(sorted_month_strs):
            demand_col = f"{ym}需求"
            alloc_col = f"{ym}分配"
            demand = pd.to_numeric(row.get(demand_col, 0), errors="coerce") or 0.0

            if i == 0:
                # 第一个月
                wo_before = sum(
                    pd.to_numeric(row.get(col, 0), errors="coerce") or 0
                    for col, wo_date in wo_cols if wo_date < first_date
                )

                total_available = (
                    pd.to_numeric(row.get("分片晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("工程晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("已测晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("未测晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("CP在制（Total）", 0), errors="coerce") +
                    pd.to_numeric(row.get("Fab warehouse", 0), errors="coerce") * wafer_unit +
                    wo_before * wafer_unit
                )
            else:
                prev_month = sorted_month_strs[i - 1]
                wo_col = f"{prev_month} WO"
                wo = pd.to_numeric(row.get(wo_col, 0), errors="coerce") or 0
                total_available = rest_prev + wo * wafer_unit

            delta = total_available - demand
            allocated = demand if delta > 0 else total_available
            rest_prev = max(delta, 0)
            df.at[idx, alloc_col] = round(allocated, 3)

    return df


def append_cumulative_gap_columns(df_unique_wafer: pd.DataFrame, start_date) -> pd.DataFrame:
    """
    为 df_unique_wafer 增加“YYYY-MM累积缺口”列。
    计算方式：
        累积缺口 = 截至当前月所有“YYYY-MM需求”总和 - InvPart - total_available

    参数:
        df_unique_wafer: 包含各项仓库存、需求、WO、InvPart 等信息
        start_date: 起始日期，datetime 类型
    返回:
        含新增“YYYY-MM累积缺口”列的 DataFrame
    """
    first_date = datetime(start_date.year, start_date.month, 1)
    df = df_unique_wafer.copy()

    # 匹配“YYYY-MM需求”列
    pattern = re.compile(r"^(\d{4})-(\d{2})需求$")
    demand_cols = [col for col in df.columns if pattern.match(col)]
    if not demand_cols:
        raise ValueError("❌ 未找到任何“x月需求”列")

    # 提取和排序月份
    month_keys = [(col, pattern.match(col).group(1) + "-" + pattern.match(col).group(2)) for col in demand_cols]
    month_keys.sort(key=lambda x: x[1])
    sorted_demand_cols = [col for col, _ in month_keys]
    sorted_months = [month for _, month in month_keys]
    gap_cols = [f"{m}累积缺口" for m in sorted_months]

    for col in gap_cols:
        df[col] = 0.0

    # 获取所有 WO 列及其日期
    wo_pattern = re.compile(r"^(\d{4})-(\d{2}) WO$")
    wo_cols = []
    for col in df.columns:
        match = wo_pattern.match(str(col))
        if match:
            y, m = int(match.group(1)), int(match.group(2))
            wo_cols.append((col, datetime(y, m, 1)))
    wo_cols.sort(key=lambda x: x[1])

    # 逐行计算
    for idx, row in df.iterrows():
        wafer_unit = pd.to_numeric(row.get("单片数量", 1.0), errors="coerce") or 1.0
        total_available = 0

        for i, month in enumerate(sorted_months):
            demand_sum = sum(
                pd.to_numeric(row.get(col, 0), errors="coerce") or 0.0
                for col in sorted_demand_cols[:i + 1]
            )

            inv_part = pd.to_numeric(row.get("InvPart", 0), errors="coerce") or 0.0
            gap_col = f"{month}累积缺口"

            if i == 0:
                wo_before_sum = sum(
                    pd.to_numeric(row.get(col, 0), errors="coerce") or 0
                    for col, wo_date in wo_cols if wo_date < first_date
                )
                total_available = (
                    pd.to_numeric(row.get("分片晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("工程晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("已测晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("未测晶圆仓", 0), errors="coerce") +
                    pd.to_numeric(row.get("CP在制（Total）", 0), errors="coerce") +
                    pd.to_numeric(row.get("Fab warehouse", 0), errors="coerce") * wafer_unit +
                    wo_before_sum * wafer_unit +
                    inv_part
                )
            else:
                prev_month = sorted_months[i - 1]
                wo_col = f"{prev_month} WO"
                wo = pd.to_numeric(row.get(wo_col, 0), errors="coerce") or 0
                total_available += wo * wafer_unit

            gap = (demand_sum - total_available) / wafer_unit
            df.at[idx, gap_col] = round(gap, 3)

    return df

def merge_cumulative_gap_header(ws, df):
    """
    合并“yyyy-mm累积缺口”列第一行单元格，写入标题“晶圆缺口累加（片）”
    """
    pattern = re.compile(r"^20\d{2}-\d{2}累积缺口$")
    gap_cols = [col for col in df.columns if pattern.match(str(col))]
    if not gap_cols:
        return

    start_col = df.columns.get_loc(gap_cols[0]) + 1
    end_col = df.columns.get_loc(gap_cols[-1]) + 1

    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)

    merge_range = f"{start_letter}1:{end_letter}1"
    ws.merge_cells(merge_range)

    cell = ws.cell(row=1, column=start_col)
    cell.value = "晶圆缺口累加（片）"
    cell.alignment = Alignment(horizontal="center", vertical="center")
