import re
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import numbers
from sheet_add import clean_df

def init_monthly_fields(main_plan_df: pd.DataFrame, start_date: datetime = None) -> tuple[pd.DataFrame, list[str]]:
    """
    自动识别主计划中预测字段（如“2025-10预测”），添加 HEADER_TEMPLATE 中所有相关字段。
    初始化为空字符串 ""，并保留 start_date 对齐或之后的月份。

    返回：
    - 修改后的 main_plan_df（带新列）
    - forecast_months: 所有识别出的月份列表（如 “2025-10”）
    """
    HEADER_TEMPLATE = [
        "销售数量", "销售金额", "成品投单计划", "半成品投单计划", "投单计划调整",
        "成品可行投单", "半成品可行投单", "成品实际投单", "半成品实际投单",
        "回货计划", "回货计划调整", "PC回货计划", "回货实际"
    ]

    # 提取预测字段中符合格式的 YYYY-MM 字符串
    month_pattern = re.compile(r"^(\d{4}-\d{2})预测$")
    all_months = sorted({
        match.group(1) for col in main_plan_df.columns
        if isinstance(col, str) and (match := month_pattern.match(col.strip()))
    })

    if not all_months:
        return main_plan_df, []

    # 将 string 转为 Timestamp，并筛选大于等于 start_date 的月份
    base_date = pd.Timestamp(start_date.replace(day=1)) if start_date else pd.Timestamp.today().replace(day=1)
    forecast_months = [m for m in all_months if pd.Timestamp(m + "-01") >= base_date]

    df = main_plan_df.copy()
    for ym in forecast_months:
        for header in HEADER_TEMPLATE:
            col = f"{ym}{header}"
            if col not in df.columns:
                df[col] = ""

    return df, forecast_months

    
def safe_col(df: pd.DataFrame, col: str) -> pd.Series:
    """确保列为数字，若不存在则返回 0"""
    return pd.to_numeric(df[col], errors="coerce").fillna(0) if col in df.columns else pd.Series(0, index=df.index)

def generate_monthly_fg_plan(main_plan_df: pd.DataFrame, forecast_months: list[str]) -> pd.DataFrame:
    """
    生成每月“成品投单计划”列，规则：
    - 第一个月：InvPart + max(预测, 未交) + max(预测, 未交)（下月） - 成品仓 - 成品在制
    - 后续月份：max(预测, 未交)（下月） + （上月投单 - 上月实际投单）

    参数：
    - main_plan_df: 主计划表（含所有字段）
    - forecast_months: 所有月份的列表（如 ['2025-07', '2025-08', ...]）

    返回：
    - main_plan_df: 添加了成品投单计划字段的 DataFrame
    """
    df_plan = pd.DataFrame(index=main_plan_df.index)

    for idx, ym in enumerate(forecast_months[:-1]):  # 最后一个月不生成
        this_month = ym
        next_month = forecast_months[idx + 1]
        prev_month = forecast_months[idx - 1] if idx > 0 else None

        # 构造字段名
        col_forecast_this = f"{this_month}预测"
        col_order_this = f"未交订单 {this_month}"
        col_forecast_next = f"{next_month}预测"
        col_order_next = f"未交订单 {next_month}"
        col_target = f"{this_month}成品投单计划"
        col_actual_prod = f"{prev_month}成品实际投单" if prev_month else None
        col_target_prev = f"{prev_month}成品投单计划" if prev_month else None
        col_sales_this = f"{this_month}销售数量"

        # 安全提取列，如果缺失则填 0
        def get(col):
            return pd.to_numeric(main_plan_df[col], errors="coerce").fillna(0) if col in main_plan_df.columns else pd.Series(0, index=main_plan_df.index)

        def get_plan(col):
            return pd.to_numeric(df_plan[col], errors="coerce").fillna(0) if col in df_plan.columns else pd.Series(0, index=main_plan_df.index)

        if idx == 0:
            cond = (get(col_order_this) + get(col_sales_this) > get(col_forecast_this))
            for row_idx in main_plan_df.index:
                v_invpart = get("InvPart").at[row_idx]
                v_fg_inv = get("成品仓").at[row_idx]
                v_fg_wip = get("成品在制").at[row_idx]
                v_order_this = get(col_order_this).at[row_idx]
                v_sales_this = get(col_sales_this).at[row_idx]
                v_forecast_this = get(col_forecast_this).at[row_idx]
                v_forecast_next = get(col_forecast_next).at[row_idx]
                v_order_next = get(col_order_next).at[row_idx]
                max_next = max(v_forecast_next, v_order_next)

                if cond.at[row_idx]:
                    result = v_invpart + v_order_this + max_next - v_fg_inv - v_fg_wip
                else:
                    result = v_invpart + v_forecast_this - v_sales_this + max_next - v_fg_inv - v_fg_wip

                df_plan.at[row_idx, col_target] = result
        else:
            for row_idx in main_plan_df.index:
                v_prev_plan = get_plan(col_target_prev).at[row_idx]
                v_actual = get(col_actual_prod).at[row_idx] if col_actual_prod else 0
                v_forecast_next = get(col_forecast_next).at[row_idx]
                v_order_next = get(col_order_next).at[row_idx]
                result = v_prev_plan - v_actual + max(v_forecast_next, v_order_next)
                df_plan.at[row_idx, col_target] = result

    # ✅ 若主计划中没有这些列，先批量添加空列（防止碎片化 + 保证后续写入成功）
    for col in df_plan.columns:
        if col not in main_plan_df.columns:
            main_plan_df[col] = ""
    
    # ✅ 写入 df_plan 内容（成品投单计划）
    for col in df_plan.columns:
        main_plan_df[col] = df_plan[col]


    return main_plan_df


def generate_monthly_semi_plan(main_plan_df: pd.DataFrame, forecast_months: list[str],
                                mapping_df: pd.DataFrame) -> pd.DataFrame:
    """
    生成每月“半成品投单计划”列，规则：
    - 第一个月：成品投单计划 - 半成品在制 - 半成品仓
    - 后续月份：上月半成品投单计划 - 上月半成品实际投单 + max(预测, 未交订单)

    参数：
    - main_plan_df: 主计划表（含所有字段）
    - forecast_months: 所有月份的列表（如 ["2025-07", "2025-08", ...]）
    - mapping_df: 新旧料号及半成品映射表，含“新品名”和“半成品”列

    返回：
    - main_plan_df: 添加了半成品投单计划字段的 DataFrame
    """
    tmp = mapping_df[["新品名", "半成品"]].copy()
    tmp = clean_df(tmp)
    tmp = tmp[tmp["半成品"].notna() & (tmp["半成品"].astype(str).str.strip() != "")]
    
    combined_names = pd.Series(
        tmp["新品名"].astype(str).str.strip().tolist() +
        tmp["半成品"].astype(str).str.strip().tolist()
    ).dropna().unique().tolist()

    mask = main_plan_df["品名"].astype(str).str.strip().isin(combined_names)

    df_plan = pd.DataFrame(index=main_plan_df.index)

    for idx, ym in enumerate(forecast_months[:-1]):  # 最后一个月不生成
        this_month = ym
        next_month = forecast_months[idx + 1]
        prev_month = forecast_months[idx - 1] if idx > 0 else None

        # 构造字段名
        col_target = f"{this_month}半成品投单计划"
        col_plan_this = f"{this_month}成品投单计划"
        col_forecast_next = f"{next_month}预测"
        col_order_next = f"未交订单 {next_month}"
        col_actual_prod = f"{prev_month}半成品实际投单" if prev_month else None
        col_target_prev = f"{prev_month}半成品投单计划" if prev_month else None

        # 安全提取列
        def get(col):
            return pd.to_numeric(main_plan_df[col], errors="coerce").fillna(0) if col in main_plan_df.columns else pd.Series(0, index=main_plan_df.index)

        def get_plan(col):
            return pd.to_numeric(df_plan[col], errors="coerce").fillna(0) if col in df_plan.columns else pd.Series(0, index=main_plan_df.index)

        if idx == 0:
            plan_this = get(col_plan_this)
            sfg = get("半成品仓")
            sfg_wip = get("半成品在制")
            result = plan_this - sfg - sfg_wip
            df_plan[col_target] = result
        else:
            for row_idx in main_plan_df.index:
                prev_plan = get_plan(col_target_prev).at[row_idx]
                actual_prod = get(col_actual_prod).at[row_idx] if col_actual_prod else 0
                forecast_next = get(col_forecast_next).at[row_idx]
                order_next = get(col_order_next).at[row_idx]
                result = prev_plan - actual_prod + max(forecast_next, order_next)
                df_plan.at[row_idx, col_target] = result

    # ✅ 若主计划中没有这些列，先批量添加空列
    for col in df_plan.columns:
        if col not in main_plan_df.columns:
            main_plan_df[col] = ""
    
    # ✅ 将结果回填（保留原有 mask 过滤）
    for col in df_plan.columns:
        main_plan_df.loc[mask, col] = df_plan.loc[mask, col]

    
    return main_plan_df


def aggregate_actual_fg_orders(main_plan_df: pd.DataFrame,
                               df_order: pd.DataFrame,
                               forecast_months: list[str]) -> pd.DataFrame:
    """
    抓取“成品实际投单”并写入 main_plan_df，每月写入“2025-10成品实际投单”等列。
    """
    if df_order.empty or not forecast_months:
        return main_plan_df

    df_order = df_order[["下单日期", "回货明细_回货品名", "回货明细_回货数量"]].dropna()
    df_order["回货明细_回货品名"] = df_order["回货明细_回货品名"].astype(str).str.strip()
    df_order["下单年月"] = pd.to_datetime(df_order["下单日期"], errors="coerce").dt.strftime("%Y-%m")

    main_plan_df = main_plan_df.copy()
    main_plan_df["品名"] = main_plan_df["品名"].astype(str).str.strip()

    # 初始化并强制成数值列
    for ym in forecast_months:
        col = f"{ym}成品实际投单"
        if col not in main_plan_df.columns:
            main_plan_df[col] = 0
        else:
            main_plan_df[col] = pd.to_numeric(main_plan_df[col], errors='coerce').fillna(0)

    # 遍历订单写入主计划
    for _, row in df_order.iterrows():
        part = row["回货明细_回货品名"]
        qty = row["回货明细_回货数量"]
        ym = row["下单年月"]
        col_name = f"{ym}成品实际投单"

        if col_name in main_plan_df.columns:
            main_plan_df.loc[main_plan_df["品名"] == part, col_name] += qty

    return main_plan_df


def aggregate_actual_sfg_orders(main_plan_df: pd.DataFrame,
                                df_order: pd.DataFrame,
                                mapping_df: pd.DataFrame,
                                forecast_months: list[str],
                                debug: bool = False) -> pd.DataFrame:
    """
    提取“半成品实际投单”并写入主计划表，包括新品名和半成品本身行。
    列格式为 “2025-10半成品实际投单”。
    """
    if df_order.empty or mapping_df.empty or not forecast_months:
        return main_plan_df

    df_order = df_order[["下单日期", "回货明细_回货品名", "回货明细_回货数量"]].dropna()
    df_order["回货明细_回货品名"] = df_order["回货明细_回货品名"].astype(str).str.strip()
    df_order["下单年月"] = pd.to_datetime(df_order["下单日期"], errors="coerce").dt.strftime("%Y-%m")

    main_plan_df = main_plan_df.copy()
    main_plan_df["品名"] = main_plan_df["品名"].astype(str).str.strip()

    semi_mapping = mapping_df[mapping_df["半成品"].notna() & (mapping_df["半成品"] != "")]
    semi_dict = dict(zip(
        semi_mapping["半成品"].astype(str).str.strip(),
        semi_mapping["新品名"].astype(str).str.strip()
    ))

    # 初始化列并强制为数值
    for ym in forecast_months:
        col = f"{ym}半成品实际投单"
        if col not in main_plan_df.columns:
            main_plan_df[col] = 0
        main_plan_df[col] = pd.to_numeric(main_plan_df[col], errors='coerce').fillna(0)

    # 遍历订单
    for _, row in df_order.iterrows():
        part = row["回货明细_回货品名"]
        qty = row["回货明细_回货数量"]
        ym = row["下单年月"]
        col_name = f"{ym}半成品实际投单"

        if ym not in forecast_months or part not in semi_dict:
            continue

        new_part = semi_dict[part]

        # 写入新品名
        if new_part in main_plan_df["品名"].values:
            main_plan_df.loc[main_plan_df["品名"] == new_part, col_name] += qty
            if debug:
                print(f"✅ 写入新品名 {new_part} → {col_name} += {qty}")
        else:
            if debug:
                print(f"⚠️ 未找到新品名 {new_part}，跳过")

        # 写入半成品自身
        if part in main_plan_df["品名"].values:
            main_plan_df.loc[main_plan_df["品名"] == part, col_name] += qty
            if debug:
                print(f"✅ 同时写入半成品自身 {part} → {col_name} += {qty}")
        else:
            if debug:
                print(f"⚠️ 半成品 {part} 本身不在主计划中")

    return main_plan_df




def aggregate_actual_arrivals(main_plan_df: pd.DataFrame, df_arrival: pd.DataFrame, forecast_months: list[str]) -> pd.DataFrame:
    """
    从“到货明细”中提取回货实际数量并填入主计划表。
    
    参数：
    - main_plan_df: 主计划 DataFrame（需包含“品名”列）
    - df_arrival: 到货明细 DataFrame，含“到货日期”、“品名”、“允收数量”
    - forecast_months: 月份字符串列表，如 ["2025-07", "2025-08"]
    
    返回：
    - main_plan_df: 添加了“yyyy-mm回货实际”的列
    """
    if df_arrival.empty or not forecast_months:
        return main_plan_df

    # 保留有效列并清洗
    df_arrival = df_arrival[["到货日期", "品名", "允收数量"]].dropna()
    df_arrival["品名"] = df_arrival["品名"].astype(str).str.strip()
    df_arrival["到货月份"] = pd.to_datetime(df_arrival["到货日期"], errors="coerce").dt.strftime("%Y-%m")

    # 初始化结果表
    result_df = pd.DataFrame({"品名": main_plan_df["品名"].astype(str)})
    for ym in forecast_months:
        result_df[f"{ym}回货实际"] = 0

    # 汇总每月数据
    for _, row in df_arrival.iterrows():
        part = row["品名"]
        qty = row["允收数量"]
        ym = row["到货月份"]
        col = f"{ym}回货实际"
        if ym in forecast_months:
            match_idx = result_df[result_df["品名"] == part].index
            if not match_idx.empty:
                result_df.loc[match_idx[0], col] += qty

    # 写入主计划表
    for col in result_df.columns[1:]:
        main_plan_df[col] = result_df[col]

    return main_plan_df

def aggregate_sales_quantity_and_amount(main_plan_df: pd.DataFrame, df_sales: pd.DataFrame, forecast_months: list[str]) -> pd.DataFrame:
    """
    将销货明细中的销售数量和销售金额按照 yyyy-mm 格式月份填入主计划表。

    参数：
    - main_plan_df: 主计划 DataFrame（含“品名”列）
    - df_sales: 销货明细 DataFrame，含“交易日期”、“品名”、“数量”、“原币金额”
    - forecast_months: 月份字符串列表，如 ["2025-07", "2025-08"]

    返回：
    - main_plan_df: 添加了“yyyy-mm销售数量”和“yyyy-mm销售金额”的列
    """
    if df_sales.empty or not forecast_months:
        return main_plan_df

    df_sales = df_sales[["交易日期", "品名", "数量", "原币金额"]].dropna()
    df_sales["品名"] = df_sales["品名"].astype(str).str.strip()
    df_sales["销售月份"] = pd.to_datetime(df_sales["交易日期"], errors="coerce").dt.strftime("%Y-%m")

    result_qty = pd.DataFrame({"品名": main_plan_df["品名"].astype(str)})
    result_amt = pd.DataFrame({"品名": main_plan_df["品名"].astype(str)})

    for ym in forecast_months:
        result_qty[f"{ym}销售数量"] = 0
        result_amt[f"{ym}销售金额"] = 0

    for _, row in df_sales.iterrows():
        part = row["品名"]
        qty = row["数量"]
        amt = row["原币金额"]
        ym = row["销售月份"]
        if ym in forecast_months:
            col_qty = f"{ym}销售数量"
            col_amt = f"{ym}销售金额"
            match_idx = result_qty[result_qty["品名"] == part].index
            if not match_idx.empty:
                result_qty.loc[match_idx[0], col_qty] += qty
                result_amt.loc[match_idx[0], col_amt] += amt

    for col in result_qty.columns[1:]:
        main_plan_df[col] = result_qty[col]
    for col in result_amt.columns[1:]:
        main_plan_df[col] = result_amt[col]

    return main_plan_df


def generate_monthly_adjust_plan(main_plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    根据已有字段直接填充投单计划调整列。
    第一个月为空，后续为公式字符串。
    """
    adjust_cols = [col for col in main_plan_df.columns if "投单计划调整" in col]
    fg_plan_cols = [col for col in main_plan_df.columns if "成品投单计划" in col and "半成品" not in col]
    fg_actual_cols = [col for col in main_plan_df.columns if "成品实际投单" in col and "半成品" not in col]

    if not adjust_cols or not fg_plan_cols or not fg_actual_cols:
        raise ValueError("❌ 缺少必要的列：投单计划调整 / 成品投单计划 / 成品实际投单")

    for i, col in enumerate(adjust_cols):
        if i == 0:
            # 第一个月为空字符串
            main_plan_df[col] = ""
        else:
            # 后续月：写入公式
            curr_plan_col = fg_plan_cols[i] if i < len(fg_plan_cols) else None
            prev_plan_col = fg_plan_cols[i - 1]
            prev_actual_col = fg_actual_cols[i - 1]

            # 获取 Excel 的列号（+1 因为 openpyxl 是从 1 开始）
            col_curr_plan = get_column_letter(main_plan_df.columns.get_loc(curr_plan_col) + 1)
            col_prev_plan = get_column_letter(main_plan_df.columns.get_loc(prev_plan_col) + 1)
            col_prev_actual = get_column_letter(main_plan_df.columns.get_loc(prev_actual_col) + 1)

            def build_formula(row_idx: int) -> str:
                row_num = row_idx + 3  # 数据起始于 Excel 第 3 行
                return f"={col_curr_plan}{row_num}+({col_prev_plan}{row_num}-{col_prev_actual}{row_num})"

            main_plan_df[col] = [build_formula(i) for i in range(len(main_plan_df))]
    return main_plan_df

def generate_monthly_return_adjustment(main_plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    填写“回货计划调整”列：
    - 第一个月为空
    - 后续月份：= 本月回货计划 + (上月成品实际投单 - 上月投单计划调整)
    """
    adjust_return_cols = [col for col in main_plan_df.columns if "回货计划调整" in col]
    return_plan_cols = [col for col in main_plan_df.columns if "回货计划" in col and "调整" not in col and "PC" not in col]
    actual_plan_cols = [col for col in main_plan_df.columns if "成品实际投单" in col and "半成品" not in col]
    adjust_plan_cols = [col for col in main_plan_df.columns if "投单计划调整" in col]

    for i in range(len(adjust_return_cols)):
        col_adjust = adjust_return_cols[i]
        col_return = return_plan_cols[i]

        # 本月列索引
        col_idx_return = main_plan_df.columns.get_loc(col_return) + 1
        col_idx_adjust = main_plan_df.columns.get_loc(col_adjust) + 1

        # 上月列（用于差值计算）
        if i > 0:
            col_idx_prev_actual = main_plan_df.columns.get_loc(actual_plan_cols[i - 1]) + 1
            col_idx_prev_adjust = main_plan_df.columns.get_loc(adjust_plan_cols[i - 1]) + 1

        for row in range(3, len(main_plan_df) + 3):  # 第3行起是数据行
            if i == 0:
                main_plan_df.at[row - 3, col_adjust] = ""
            else:
                # 本月回货计划
                col_r = get_column_letter(col_idx_return)
                # 上月：成品实际投单与投单计划调整
                col_prev_actual = get_column_letter(col_idx_prev_actual)
                col_prev_adjust = get_column_letter(col_idx_prev_adjust)

                formula = f"={col_r}{row} + ({col_prev_actual}{row} - {col_prev_adjust}{row})"
                main_plan_df.at[row - 3, col_adjust] = formula

    return main_plan_df

def generate_monthly_return_plan(main_plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    回货计划填写逻辑：
    - 第一个月为空；
    - 从第二个月开始，等于同一行第当前列前18列的值（通过公式表示）。
    """
    # 找出所有“回货计划”列（不含“调整”）
    return_plan_cols = [col for col in main_plan_df.columns if "回货计划" in col and "调整" not in col and "PC" not in col]
    
    # 处理每一个回货计划列
    for i, col in enumerate(return_plan_cols):
        if i == 0:
            # 第一个月为空
            main_plan_df[col] = ""
        else:
            # 获取该列在 DataFrame 中的位置
            col_idx = main_plan_df.columns.get_loc(col)
            prev_18_idx = col_idx - 18
            if prev_18_idx < 0:
                raise ValueError(f"❌ 第{i+1}月回货计划前18列不存在，列索引越界。")

            # 获取引用的前18列名
            ref_col = main_plan_df.columns[prev_18_idx]

            # 构造 Excel 公式：=INDIRECT(ADDRESS(ROW(), col_index))
            col_letter = get_column_letter(prev_18_idx + 1)  # Excel 列号从 1 开始
            main_plan_df[col] = f"={col_letter}" + (main_plan_df.index + 3).astype(str)

    return main_plan_df
    
def format_monthly_grouped_headers(ws):
    """
    处理格式为“2025-12成品投单计划”的字段：
    - 从识别到的起始列开始，每13列为一个月份组
    - 第1行合并写 “2025-12”
    - 第2行去掉前缀“2025-12”，仅保留字段名
    - 每月块使用不同颜色填充前两行
    """
    from openpyxl.styles import PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter
    import re

    # 查找第2行中第一个“销售数量”字段的列号作为起点
    start_col = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col_idx).value
        if isinstance(cell_value, str) and "销售数量" in cell_value:
            start_col = col_idx
            break
    if start_col is None:
        raise ValueError("❌ 未在第2行找到“销售数量”字段，无法定位起始列")

    row_1, row_2 = 1, 2
    max_col = ws.max_column

    fill_colors = [
        "FFF2CC", "D9EAD3", "D0E0E3", "F4CCCC", "EAD1DC", "CFE2F3", "FFE599"
    ]

    month_pattern = re.compile(r"^(\d{4}-\d{2})(.+)$")
    col = start_col
    month_index = 0

    while col <= max_col:
        month_title = None
        fill_color = PatternFill(start_color=fill_colors[month_index % len(fill_colors)],
                                 end_color=fill_colors[month_index % len(fill_colors)],
                                 fill_type="solid")

        for offset in range(13):
            curr_col = col + offset
            if curr_col > max_col:
                break
            cell = ws.cell(row=row_2, column=curr_col)
            value = cell.value
            if isinstance(value, str):
                match = month_pattern.match(value.strip())
                if match:
                    if month_title is None:
                        month_title = match.group(1)
                    # 更新字段名：只保留后缀，如“成品投单计划”
                    cell.value = match.group(2)

            # 应用颜色到第2行
            ws.cell(row=row_2, column=curr_col).fill = fill_color

        if month_title:
            start_letter = get_column_letter(col)
            end_letter = get_column_letter(min(col + 12, max_col))
            ws.merge_cells(f"{start_letter}{row_1}:{end_letter}{row_1}")
            top_cell = ws.cell(row=row_1, column=col)
            top_cell.value = month_title
            top_cell.alignment = Alignment(horizontal="center", vertical="center")
            top_cell.font = Font(bold=True)
            top_cell.fill = fill_color

        col += 13
        month_index += 1


def highlight_production_plan_cells(ws, df):
    """
    根据规则给所有“成品投单计划”列标色：
    - < 0：红色
    - < 安全库存：黄色
    - > 2 * 安全库存：橙色
    """
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # 获取列位置
    plan_cols = [col for col in df.columns if "成品投单计划" in col and "半成品" not in col]
    safety_col = "InvPart"
    if safety_col not in df.columns:
        raise ValueError("❌ 缺少“安全库存”列，无法对成品投单计划进行标色。")

    for col in plan_cols:
        col_idx = df.columns.get_loc(col) + 1  # openpyxl是1-based
        for i, val in enumerate(df[col]):
            row_idx = i + 3  # 因为第1行是合并标题，第2行是字段名
            safety = df.at[i, safety_col]

            # 进行数值判断（确保为float）
            try:
                val = float(val)
                safety = float(safety)
            except:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            if val < 0:
                cell.fill = red_fill
            elif val < safety:
                cell.fill = yellow_fill
            elif val > 2 * safety:
                cell.fill = orange_fill

def drop_last_forecast_month_columns(main_plan_df: pd.DataFrame, forecast_months: list[int]) -> pd.DataFrame:
    """
    删除 AC列（第29列）后所有含有最后一个预测月份的字段列，如 '12月销售数量' 等。
    """
    if not forecast_months:
        return main_plan_df  # 无预测月份，不处理

    last_valid_month = forecast_months[-1]
    last_month_str = f"{last_valid_month}月"

    # 起始列为 AC = 第29列，0-based index 为 28
    fixed_part = main_plan_df.iloc[:, :28]
    dynamic_part = main_plan_df.iloc[:, 28:]

    # 仅保留不包含最后预测月的列
    dynamic_part = dynamic_part.loc[:, ~dynamic_part.columns.str.contains(fr"^{last_month_str}")]

    # 合并回主表
    cleaned_df = pd.concat([fixed_part, dynamic_part], axis=1)

    return cleaned_df

