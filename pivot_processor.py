import re
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime, date
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import FILE_KEYWORDS, FIELD_MAPPINGS, pivot_config, RENAME_MAP
from excel_utils import (
    adjust_column_width, 
    highlight_replaced_names_in_main_sheet, 
    reorder_main_plan_by_unfulfilled_sheet, 
    format_currency_columns_rmb,
    format_thousands_separator,
    add_sheet_hyperlinks
)
from mapping_utils import (
    clean_mapping_headers, 
    replace_all_names_with_mapping, 
    apply_mapping_and_merge, 
    apply_extended_substitute_mapping,
    apply_all_name_replacements,
    extract_mappings
)
from data_utils import (
    fill_spec_and_wafer_info, 
    fill_packaging_info
)
from summary import (
    merge_safety_inventory,
    merge_safety_header,
    append_unfulfilled_summary_columns_by_date,
    merge_unfulfilled_order_header, 
    append_forecast_to_summary,
    merge_forecast_header,
    merge_finished_inventory_with_warehouse_types,
    merge_inventory_header,
    append_product_in_progress,
    merge_product_in_progress_header,
    append_order_delivery_amount_columns,
    merge_order_delivery_amount,
    append_forecast_accuracy_column,
    merge_forecast_accuracy
)
from production_plan import (
    init_monthly_fields,
    generate_monthly_fg_plan,
    aggregate_actual_fg_orders,
    aggregate_actual_sfg_orders,
    aggregate_actual_arrivals,
    aggregate_sales_quantity_and_amount,
    generate_monthly_semi_plan,
    generate_monthly_adjust_plan,
    generate_monthly_return_adjustment,
    generate_monthly_return_plan,
    format_monthly_grouped_headers,
    highlight_production_plan_cells,
    drop_last_forecast_month_columns
)
from sheet_add import clean_df, append_all_standardized_sheets, append_original_cp_sheets
from pivot_generator import generate_monthly_pivots, standardize_uploaded_keys
from cp_file_utils import merge_cp_files_by_keyword, generate_fab_summary, format_fab_summary_month_headers
from wafer_utils import(
    extract_wafer_with_grossdie_raw, 
    append_inventory_columns, 
    append_wafer_inventory_by_warehouse,
    merge_wafer_inventory_columns,
    append_cp_wip_total,
    merge_cp_wip_column,
    append_fab_warehouse_quantity,
    merge_fab_warehouse_column,
    append_monthly_wo_from_weekly_fab,
    merge_monthly_fab_wo_columns,
    append_monthly_demand_from_fg_plan,
    merge_fg_plan_columns,
    allocate_fg_demand_monthly,
    fill_columns_c_and_right_with_zero,
    merge_allocation_header,
    append_monthly_gap_columns,
    merge_monthly_gap_columns,
    append_cumulative_gap_columns,
    merge_cumulative_gap_header
)

class PivotProcessor:
    def process(self, uploaded_files: dict, uploaded_cp_files: dict, df_grossdie, output_buffer, additional_sheets: dict = None, start_date: date = None):
        """
        替换品名、新建主计划表，并直接写入 Excel 文件（含列宽调整、标题行）。
        """
        # === 标准化上传文件名 ===
        # 成品文件
        self.dataframes = {}
        for filename, file_obj in uploaded_files.items():
            matched = False
            for keyword, standard_name in FILE_KEYWORDS.items():
                if keyword in filename:
                    self.dataframes[standard_name] = pd.read_excel(file_obj)
                    matched = True
                    break
            if not matched:
                st.warning(f"⚠️ 上传文件 `{filename}` 未识别关键词，跳过")

        # 晶圆文件
        self.cp_dataframes = {}
        self.SH_fabout = {}
        cp_keywords = ["华虹", "先进", "DB", "上华1厂", "上华2厂", "上华5厂"]
        cp_file_counter = {k: 0 for k in cp_keywords}
        
        for filename, file_obj in uploaded_cp_files.items():
            matched = False
            for keyword in cp_keywords:
                if keyword in filename:
                    cp_file_counter[keyword] += 1
                    suffix = str(cp_file_counter[keyword])
                    new_key = f"{keyword}{suffix}" if cp_file_counter[keyword] > 1 else keyword
                    if keyword == "DB":
                        self.cp_dataframes[new_key] = pd.read_excel(file_obj, header=1)
                    elif keyword.startswith("上华"):
                        self.cp_dataframes[new_key] = pd.read_excel(file_obj, sheet_name = "wip")
                        self.SH_fabout[new_key] = pd.read_excel(file_obj, sheet_name = "fabout")
                    else:
                        self.cp_dataframes[new_key] = pd.read_excel(file_obj)
                    matched = True
                    break
            if not matched:
                st.warning(f"⚠️ CP 文件 `{filename}` 未包含关键字，已跳过")

        self.cp_dataframes = merge_cp_files_by_keyword(self.cp_dataframes)
        self.SH_fabout = merge_cp_files_by_keyword(self.SH_fabout)

        # === 标准化新旧料号表 ===
        self.additional_sheets = additional_sheets
        mapping_df = self.additional_sheets.get("赛卓-新旧料号")
        if mapping_df is None or mapping_df.empty:
            raise ValueError("❌ 缺少新旧料号映射表，无法进行品名替换。")

        mapping_new, mapping_semi, mapping_sub = extract_mappings(mapping_df)
        
        # === 构建主计划 ===
        headers = ["晶圆品名", "规格", "品名", "封装厂", "封装形式", "PC"]
        main_plan_df = pd.DataFrame(columns=headers)

        ## == 品名 ==
        df_unfulfilled = self.dataframes.get("赛卓-未交订单")
        df_forecast = self.additional_sheets.get("赛卓-预测")

        name_unfulfilled = []
        name_forecast = []

        if df_unfulfilled is not None and not df_unfulfilled.empty:
            col_name = FIELD_MAPPINGS["赛卓-未交订单"]["品名"]
            name_unfulfilled = df_unfulfilled[col_name].astype(str).str.strip().tolist()

        if df_forecast is not None and not df_forecast.empty:
            col_name = FIELD_MAPPINGS["赛卓-预测"]["品名"]
            name_forecast = df_forecast[col_name].astype(str).str.strip().tolist()

        all_names = pd.Series(name_unfulfilled + name_forecast)
        all_names = replace_all_names_with_mapping(all_names, mapping_new, mapping_df)
        main_plan_df = main_plan_df.reindex(index=range(len(all_names)))
        if not all_names.empty:
            main_plan_df["品名"] = all_names.values

        ## == 规格和晶圆 ==
        main_plan_df = fill_spec_and_wafer_info(
            main_plan_df,
            self.dataframes,
            self.additional_sheets,
            mapping_semi, 
            FIELD_MAPPINGS
        )

        ## == 封装厂，封装形式和PC ==
        main_plan_df = fill_packaging_info(
            main_plan_df,
            dataframes=self.dataframes,
            additional_sheets=self.additional_sheets
        )
        st.success("✅ 已合并产品信息")

        ## == 替换新旧料号、替代料号 ==
        target_sheets = [
            ("赛卓-安全库存", self.additional_sheets),
            ("赛卓-预测", self.additional_sheets),
            ("赛卓-未交订单", self.dataframes),
            ("赛卓-成品库存", self.dataframes),
            ("赛卓-成品在制", self.dataframes),
            ("赛卓-CP在制", self.dataframes),
            ("赛卓-晶圆库存", self.dataframes),
            ("赛卓-到货明细", self.dataframes),
            ("赛卓-下单明细", self.dataframes),
            ("赛卓-销货明细", self.dataframes),
        ]
        
        all_replaced_names = set()
        
        # 执行替换逻辑
        for sheet_name, container in target_sheets:
            df_new = container[sheet_name]
        
            # 主映射替换
            df_new, replaced_main = apply_mapping_and_merge(df_new, mapping_new, FIELD_MAPPINGS[sheet_name])
            all_replaced_names.update(replaced_main)
        
            # 替代映射替换（1~4）
            df_new, replaced_sub = apply_extended_substitute_mapping(df_new, mapping_sub, FIELD_MAPPINGS[sheet_name])
            all_replaced_names.update(replaced_sub)
        
            # 更新回字典
            container[sheet_name] = df_new
        
        # 最终排序
        all_replaced_names = sorted(all_replaced_names)

        ## == 安全库存 ==
        safety_df = self.additional_sheets.get("赛卓-安全库存")
        if safety_df is not None and not safety_df.empty:
            main_plan_df, unmatched_safety = merge_safety_inventory(main_plan_df, safety_df)
            st.success("✅ 已合并安全库存数据")
        
        ## == 未交订单 ==
        unfulfilled_df = self.dataframes.get("赛卓-未交订单")
        if unfulfilled_df is not None and not unfulfilled_df.empty:
            main_plan_df, unmatched_unfulfilled = append_unfulfilled_summary_columns_by_date(main_plan_df, unfulfilled_df, start_date)
            st.success("✅ 已合并未交订单数据")
        
        ## == 预测 ==
        forecast_df = self.additional_sheets.get("赛卓-预测")
        if forecast_df is not None and not forecast_df.empty:
            main_plan_df, unmatched_forecast = append_forecast_to_summary(main_plan_df, forecast_df, start_date)
            st.success("✅ 已合并预测数据")

        ## == 成品库存 ==
        finished_df = self.dataframes.get("赛卓-成品库存")
        if finished_df is not None and not finished_df.empty:
            main_plan_df, unmatched_finished = merge_finished_inventory_with_warehouse_types(main_plan_df, finished_df, mapping_semi)
            st.success("✅ 已合并成品库存数据")

        ## == 成品在制 ==
        product_in_progress_df = self.dataframes.get("赛卓-成品在制")
        if product_in_progress_df is not None and not product_in_progress_df.empty:
            main_plan_df, unmatched_in_progress = append_product_in_progress(main_plan_df, product_in_progress_df, mapping_semi)
            st.success("✅ 已合并成品在制数据")

        ## == 发货金额 ==
        if unfulfilled_df is not None and not unfulfilled_df.empty:
            main_plan_df = append_order_delivery_amount_columns(main_plan_df, unfulfilled_df, start_date)
            st.success("✅ 已合并发货金额")

        # === 投单计划 ===
        forecast_months = init_monthly_fields(main_plan_df, start_date)

        # 成品&半成品实际投单
        df_order = self.dataframes.get("赛卓-下单明细", pd.DataFrame())
        main_plan_df = aggregate_actual_fg_orders(main_plan_df, df_order, forecast_months)
        main_plan_df = aggregate_actual_sfg_orders(main_plan_df, df_order, mapping_semi, forecast_months)

        # 回货实际
        df_arrival = self.dataframes.get("赛卓-到货明细", pd.DataFrame())
        main_plan_df = aggregate_actual_arrivals(main_plan_df, df_arrival, forecast_months)

        # 销售数量&销售金额
        df_sales = self.dataframes.get("赛卓-销货明细", pd.DataFrame())
        main_plan_df = aggregate_sales_quantity_and_amount(main_plan_df, df_sales, forecast_months)

        # 成品投单计划
        main_plan_df = generate_monthly_fg_plan(main_plan_df, forecast_months)

        # 半成品投单计划
        main_plan_df = generate_monthly_semi_plan(main_plan_df, forecast_months, mapping_semi)

        # 添加预测准确率列
        main_plan_df = append_forecast_accuracy_column(main_plan_df, start_date)

        # 检查
        main_plan_df = drop_last_forecast_month_columns(main_plan_df, forecast_months)
        
        st.success("✅ 已合并投单计划")

        # === FAB_WIP_汇总 ===
        df_fab_summary = generate_fab_summary(self.cp_dataframes)
        st.success("✅ 已完成FAB_WIP_汇总")

        # === 晶圆需求汇总 ===
        ## == 单片数量 ==
        df_unique_wafer = extract_wafer_with_grossdie_raw(main_plan_df, df_grossdie)
        ## == 安全库存 ==
        df_unique_wafer = append_inventory_columns(df_unique_wafer, main_plan_df)
        ## == 晶圆库存 ==
        wafer_inventory_df = self.dataframes.get("赛卓-晶圆库存")
        df_unique_wafer = append_wafer_inventory_by_warehouse(df_unique_wafer, wafer_inventory_df)
        ## == 在制CP晶圆 ==
        df_cp_wip = self.dataframes.get("赛卓-CP在制")
        df_unique_wafer = append_cp_wip_total(df_unique_wafer, df_cp_wip)
        ## == Fabout ==
        df_unique_wafer = append_fab_warehouse_quantity(df_unique_wafer, self.SH_fabout)
        ## == Fab预计晶圆产出数量 ==
        df_unique_wafer = append_monthly_wo_from_weekly_fab(df_unique_wafer, df_fab_summary)
        ## == 成品投单需求 ==
        df_unique_wafer = append_monthly_demand_from_fg_plan(df_unique_wafer, main_plan_df)
        ## == 晶圆分配（颗） ==
        df_unique_wafer = fill_columns_c_and_right_with_zero(df_unique_wafer)
        df_unique_wafer = allocate_fg_demand_monthly(df_unique_wafer, start_date)
        ## == 晶圆缺口计算（片） ==
        df_unique_wafer = append_monthly_gap_columns(df_unique_wafer)
        ## == 晶圆缺口累加（片） ==
        df_unique_wafer = append_cumulative_gap_columns(df_unique_wafer, start_date)
        st.success("✅ 已完成晶圆需求汇总")

        # === 写入 Excel 文件（主计划）===
        timestamp = datetime.now().strftime("%Y%m%d")
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            # 获取 workbook
            wb = writer.book
            
            # === 写入Summary ===
            summary_data = [
                ["", "超链接", "备注"],
                ["数据汇总", "主计划", ""],
                ["晶圆需求汇总", "晶圆需求汇总", ""],
                ["赛卓-未交订单-汇总", "赛卓-未交订单-汇总", ""],
                ["赛卓-成品库存-汇总", "赛卓-成品库存-汇总", "关注“hold仓”“成品仓”"],
                ["赛卓-晶圆库存-汇总", "赛卓-晶圆库存-汇总", "晶圆片数已转换为对应的Die数量"],
                ["赛卓-CP在制-汇总", "赛卓-CP在制-汇总", ""],
                ["赛卓-成品在制-汇总", "赛卓-成品在制-汇总", ""],
                ["赛卓-预测", "赛卓-预测", ""],
                ["赛卓-安全库存", "赛卓-安全库存", ""],
                ["赛卓-新旧料号", "赛卓-新旧料号", ""]
            ]
            df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            # === 写入主计划 ===
            main_plan_df = clean_df(main_plan_df)
            main_plan_df.to_excel(writer, sheet_name="主计划", index=False, startrow=1)
        
            # 获取 worksheet
            ws = wb["主计划"]
        
            # 写时间戳和说明
            ws.cell(row=1, column=1, value=f"主计划生成时间：{timestamp}")            
            legend_cell = ws.cell(row=1, column=3)
            legend_cell.value = (
                "Red < 0    "
                "Yellow < 安全库存    "
                "Orange > 2 × 安全库存"
            )
            legend_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            fill = PatternFill(start_color="FFCCE6FF", end_color="FFCCE6FF", fill_type="solid")
            legend_cell.fill = fill

            # 合并单元格
            merge_safety_header(ws, main_plan_df)
            merge_unfulfilled_order_header(ws)
            merge_forecast_header(ws)
            merge_inventory_header(ws)
            merge_product_in_progress_header(ws)
            merge_order_delivery_amount(ws)
            merge_forecast_accuracy(ws)

            # 高亮显示
            format_monthly_grouped_headers(ws)
            highlight_production_plan_cells(ws, main_plan_df)
            highlight_replaced_names_in_main_sheet(ws, all_replaced_names)

            # 添加几个新 header
            new_headers = ["分析", "数量", "备注", "本月缺货", "回货计划", "回货实际", "计划1", "实际1", "计划2", "实际2", "计划3", 
                       "实际3", "计划4", "实际4", "计划5", "实际5", "差异", "分析", "7月订单可发数量", "7月订单可发金额", "全部订单可发数量", "全部订单可发金额", 
                       "结余数量", "结余金额", "单价", "7月订单价值", "回货价值", "缺货数量", "缺货价值"]
        
            
            # 从第2行开始写标题，从已有最大列+1处写入
            start_col = ws.max_column + 1
            
            for i, header in enumerate(new_headers):
                col_idx = start_col + i
                ws.cell(row=2, column=col_idx, value=header)
            
                # 可选：加粗、居中、设置高度
                cell = ws.cell(row=2, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")


            # 格式调整
            adjust_column_width(ws)
            format_currency_columns_rmb(ws)
            format_thousands_separator(ws)

            # 设置字体加粗，行高也调高一点
            bold_font = Font(bold=True)
            ws.row_dimensions[2].height = 35
    
            # 遍历这一行所有已用到的列，对单元格字体加粗、居中、垂直居中
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = bold_font
                # 垂直水平居中
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动筛选
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A2:{last_col_letter}2"
        
            # 冻结
            ws.freeze_panes = "D3"

            # === 写入晶圆需求汇总 ===
            df_unique_wafer.to_excel(writer, sheet_name="晶圆需求汇总", index=False, startrow=1)
            ws_wafer = wb["晶圆需求汇总"]
            ws_wafer.cell(row=1, column=1, value=f"主计划生成时间：{timestamp}") 
            merge_safety_header(ws_wafer, df_unique_wafer)
            merge_wafer_inventory_columns(ws_wafer, df_unique_wafer)
            merge_cp_wip_column(ws_wafer, df_unique_wafer)
            merge_fab_warehouse_column(ws_wafer, df_unique_wafer)
            merge_monthly_fab_wo_columns(ws_wafer, df_unique_wafer)
            merge_fg_plan_columns(ws_wafer, df_unique_wafer)
            merge_allocation_header(ws_wafer)
            merge_monthly_gap_columns(ws_wafer)
            merge_cumulative_gap_header(ws_wafer, df_unique_wafer)

            # 格式调整
            adjust_column_width(ws_wafer)
            format_thousands_separator(ws_wafer)
            bold_font = Font(bold=True)
            ws_wafer.row_dimensions[2].height = 35
    
            # 遍历这一行所有已用到的列，对单元格字体加粗、居中、垂直居中
            max_col = ws_wafer.max_column
            for col_idx in range(1, max_col + 1):
                cell = ws_wafer.cell(row=2, column=col_idx)
                cell.font = bold_font
                # 垂直水平居中
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动筛选
            last_col_letter = get_column_letter(ws_wafer.max_column)
            ws_wafer.auto_filter.ref = f"A2:{last_col_letter}2"
        
            # 冻结
            ws_wafer.freeze_panes = "C3"

            # === 写入 FAB_WIP_汇总 ===
            df_fab_summary.to_excel(writer, sheet_name="FAB_WIP_汇总", index=False, startrow=1)
        
            # 获取 worksheet
            ws = wb["FAB_WIP_汇总"]
            
            format_fab_summary_month_headers(ws)

            # 写时间戳和说明
            ws.cell(row=1, column=1, value=f"主计划生成时间：{timestamp}")            
    
            # 格式调整
            adjust_column_width(ws)

            # 设置字体加粗，行高也调高一点
            bold_font = Font(bold=True)
            ws.row_dimensions[2].height = 35
    
            # 遍历这一行所有已用到的列，对单元格字体加粗、居中、垂直居中
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = bold_font
                # 垂直水平居中
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动筛选
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A2:{last_col_letter}2"
        
            # 冻结
            ws.freeze_panes = "C3"
            
            # === 写入透视表 ===
            standardized_files = standardize_uploaded_keys(uploaded_files, RENAME_MAP)
            parsed_dataframes = {
                filename: pd.read_excel(file)  # 或提前 parse 完成的 DataFrame dict
                for filename, file in standardized_files.items()
            }
            pivot_tables = generate_monthly_pivots(parsed_dataframes, pivot_config)
            for sheet_name, df in pivot_tables.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
            # 写完后手动调整所有透视表 sheet 的列宽
            for sheet_name, df in pivot_tables.items():
                ws = writer.book[sheet_name]
                for col_cells in ws.columns:
                    max_length = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length * 1.2 + 10
                    ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

            # === 写入原表 ===
            df_grossdie.to_excel(writer, sheet_name="赛卓-GROSSDIE", index=False)
            append_all_standardized_sheets(writer, uploaded_files, self.additional_sheets)
            append_original_cp_sheets(writer, self.cp_dataframes)

            # === 加入超链接 ===
            ws_summary = wb["Summary"]
            add_sheet_hyperlinks(ws_summary, wb.sheetnames)
            
            for col_idx in range(1, ws_summary.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws_summary.column_dimensions[col_letter].width = 25

        output_buffer.seek(0)
       
    def set_additional_data(self, sheets_dict):
        """
        设置辅助数据表，如 预测、安全库存、新旧料号 等
        """
        self.additional_sheets = sheets_dict or {}
    
        # ✅ 对新旧料号进行列名清洗
        mapping_df = self.additional_sheets.get("赛卓-新旧料号")
        if mapping_df is not None and not mapping_df.empty:
            try:
                cleaned = clean_mapping_headers(mapping_df)
                self.additional_sheets["赛卓-新旧料号"] = cleaned
            except Exception as e:
                raise ValueError(f"❌ 新旧料号表清洗失败：{e}")

