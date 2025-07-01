import re
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

def merge_safety_inventory(summary_df: pd.DataFrame, safety_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    将安全库存表中 InvWaf 和 InvPart 信息按 '品名' 合并到汇总表中，仅根据 '品名' 匹配。
    对相同品名做 sum 汇总；未匹配的填 0。
    """
    safety_df = safety_df.rename(columns={"ProductionNO.": "品名"}).copy()
    safety_df.columns = safety_df.columns.str.strip()
    safety_df["品名"] = safety_df["品名"].astype(str).str.strip()
    safety_df["InvWaf"] = pd.to_numeric(safety_df["InvWaf"], errors="coerce").fillna(0)
    safety_df["InvPart"] = pd.to_numeric(safety_df["InvPart"], errors="coerce").fillna(0)

    safety_grouped = safety_df.groupby("品名", as_index=False)[["InvWaf", "InvPart"]].sum()

    summary_df["品名"] = summary_df["品名"].astype(str).str.strip()
    merged = summary_df.merge(safety_grouped, on="品名", how="left")

    matched_keys = set(safety_grouped["品名"])
    used_keys = set(merged[~merged[["InvWaf", "InvPart"]].isna().all(axis=1)]["品名"])
    unmatched_keys = list(matched_keys - used_keys)

    merged["InvWaf"] = merged["InvWaf"].fillna(0)
    merged["InvPart"] = merged["InvPart"].fillna(0)

    return merged, unmatched_keys

def merge_safety_header(ws: Worksheet, df: pd.DataFrame):
    """
    将“InvWaf”和“InvPart”两列的上方合并写入“安全库存”标题。
    """
    try:
        invwaf_col_idx = df.columns.get_loc("InvWaf") + 1  # openpyxl是1-indexed
        invpart_col_idx = df.columns.get_loc("InvPart") + 1

        start_col = get_column_letter(invwaf_col_idx)
        end_col = get_column_letter(invpart_col_idx)

        # 合并单元格
        ws.merge_cells(f"{start_col}1:{end_col}1")
        ws[f"{start_col}1"] = "安全库存"
        ws[f"{start_col}1"].alignment = Alignment(horizontal="center", vertical="center")
    except Exception as e:
        st.error(f"⚠️ 安全库存表头合并失败: {e}")

def append_unfulfilled_summary_columns_by_date( main_plan_df: pd.DataFrame, df_unfulfilled: pd.DataFrame,
    start_date: datetime = None ) -> tuple[pd.DataFrame, list]:
    """
    将未交订单按预交货日分为历史与未来月份，
    并将“历史未交订单”合并到第一个月的未交订单中，添加至主计划 DataFrame。
    返回合并后的主计划表和未匹配品名列表（df_unfulfilled 中存在但主计划中没有的）。
    """
    today = pd.Timestamp(start_date.replace(day=1)) if start_date else pd.Timestamp(datetime.today().replace(day=1))
    
    max_date = pd.to_datetime(df_unfulfilled["预交货日"], errors="coerce").max()
    if pd.isna(max_date):
        max_date = today
    final_month = (max_date + pd.offsets.MonthBegin(1)).replace(day=1)
    
    future_months = pd.period_range(today.to_period("M"), final_month.to_period("M"), freq="M")
    future_cols = [f"未交订单 {str(p)}" for p in future_months]

    # 复制一份未交订单，做清洗和预处理
    df = df_unfulfilled.copy()
    df["预交货日"] = pd.to_datetime(df["预交货日"], errors="coerce")
    df["未交订单数量"] = pd.to_numeric(df["未交订单数量"], errors="coerce").fillna(0)
    df["品名"] = df["品名"].astype(str).str.strip()
    # 提取年月周期
    df["月份"] = df["预交货日"].dt.to_period("M")

    # 按“品名”“月份”聚合未交订单数量
    df = df.groupby(["品名", "月份"], as_index=False)["未交订单数量"].sum()
    # 标记哪些行属于历史(月份 < 本月)
    df["是否历史"] = df["月份"] < today.to_period("M")

    # 统计每个品名的历史未交订单总量
    df_hist = (
        df[df["是否历史"]]
        .groupby("品名", as_index=False)["未交订单数量"]
        .sum()
        .rename(columns={"未交订单数量": "历史未交订单"})
    )

    # 剩余都是当月及之后月份，作为“未来”
    df_future = df[~df["是否历史"]].copy()
    # 将 Period 类型转为字符串，方便透视
    df_future["月份"] = df_future["月份"].astype(str)
    # 透视表：每个“品名”在每个月对应的未交订单数量
    df_pivot = (
        df_future
        .pivot_table(index="品名",
                     columns="月份",
                     values="未交订单数量",
                     aggfunc="sum")
        .fillna(0)
    )
    # 重命名列为“未交订单 YYYY-MM”
    df_pivot.columns = [f"未交订单 {col}" for col in df_pivot.columns]
    df_pivot = df_pivot.reset_index()

    # 确保所有 future_cols 都存在于透视结果中，缺失的列补 0
    for col in future_cols:
        if col not in df_pivot.columns:
            df_pivot[col] = 0

    # 将历史表和未来表合并在一起（outer 合并，方便计算哪些品名只有历史或只有未来）
    df_merged = pd.merge(df_hist, df_pivot, on="品名", how="outer").fillna(0)

    # 将“历史未交订单”合并到第一个月对应的“未交订单”列中
    first_col = future_cols[0]
    # 如果某品名既有历史也有未来，则把历史累加到当月
    df_merged[first_col] = df_merged[first_col] + df_merged["历史未交订单"]

    # 现在不再保留单独的“历史未交订单”列了
    df_merged = df_merged.drop(columns=["历史未交订单"])

    # 重新计算“总未交订单” = 所有 future_cols 之和
    df_merged["总未交订单"] = df_merged[future_cols].sum(axis=1)

    # 按顺序排列列: “品名”“总未交订单” + future_cols
    ordered_cols = ["品名", "总未交订单"] + future_cols
    df_merged = df_merged[ordered_cols]

    # 清洗主计划表的“品名”格式
    main_plan_df["品名"] = main_plan_df["品名"].astype(str).str.strip()
    # 将合并结果和主计划表对齐（左连接）
    result = pd.merge(main_plan_df, df_merged, on="品名", how="left")

    # 将新加入的列空值填 0
    for col in ordered_cols[1:]:
        if col in result.columns:
            result[col] = result[col].fillna(0)

    # 计算未匹配品名：df_unfulfilled 有，但主计划中没有
    all_unfulfilled_names = set(df_unfulfilled["品名"].dropna().astype(str).str.strip())
    all_main_names = set(main_plan_df["品名"].dropna().astype(str).str.strip())
    unmatched = sorted(list(all_unfulfilled_names - all_main_names))

    return result, unmatched

def merge_unfulfilled_order_header(sheet):
    """
    自动检测以“未交订单”开头的列，在第一行合并并写入“未交订单”，居中。
    
    参数:
    - sheet: openpyxl worksheet 对象
    """
    # 第2行是列名行（默认 DataFrame 用 dataframe_to_rows 写入时）
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # 找出所有“未交订单 yyyy-mm”列的索引
    unfulfilled_cols = [
        idx for idx, col in enumerate(header_row, start=1)
        if isinstance(col, str) and col.startswith("未交订单 ")
    ]

    if not unfulfilled_cols:
        return  # 没有未交订单列，不处理

    start_col = min(unfulfilled_cols) - 1
    end_col = max(unfulfilled_cols)

    # 合并单元格范围
    merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
    sheet.merge_cells(merge_range)

    # 设置合并单元格的值与居中格式
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "未交订单"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_forecast_to_summary(summary_df: pd.DataFrame, forecast_df: pd.DataFrame, 
                               start_date: datetime = None) -> tuple[pd.DataFrame, list]:
    """
    从预测表中提取当月及未来的预测信息（仅按“品名”匹配），合并至 summary_df。
    返回合并后的表格和未匹配的品名列表。

    参数:
    - summary_df: 主计划 DataFrame（需含 '品名'）
    - forecast_df: 原始预测表（需含 '生产料号' 及预测列）

    返回:
    - result: 合并后的 DataFrame
    - unmatched_keys: list[str]，未匹配的品名
    """
    today = pd.Timestamp(start_date.replace(day=1)) if start_date else pd.Timestamp(datetime.today().replace(day=1))
    this_month_int = today.month

    # ✅ 统一列名
    forecast_df = forecast_df.rename(columns={"生产料号": "品名"}).copy()
    forecast_df["品名"] = forecast_df["品名"].astype(str).str.strip()

    # ✅ 识别预测列（仅保留“x月预测”且月份 >= 当前月）    
    # 获取所有“x月预测”列，且月份合法
    month_cols = [
        col for col in forecast_df.columns
        if isinstance(col, str) and col.endswith("月预测") and "月" in col and col[:col.index("月")].isdigit()
    ]
    
    # 保留当前月及以后的预测列
    future_month_cols = [
        col for col in month_cols
        if int(col[:col.index("月")]) >= this_month_int
    ]

    if not future_month_cols:
        st.warning("⚠️ 未找到当月或未来月份的预测列（格式应为“5月预测”）")
        return summary_df, []

    # ✅ 汇总相同品名的预测值
    forecast_df[future_month_cols] = forecast_df[future_month_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    forecast_grouped = forecast_df.groupby("品名", as_index=False)[future_month_cols].sum()

    # ✅ 合并到主计划
    summary_df["品名"] = summary_df["品名"].astype(str).str.strip()
    result = summary_df.merge(forecast_grouped, on="品名", how="left")

    # ✅ 填补新预测列中的 NaN 为 0（不影响原有列）
    for col in future_month_cols:
        if col in result.columns:
            result[col] = result[col].fillna(0)

    # ✅ 找出未匹配品名
    forecast_keys = set(forecast_grouped["品名"])
    summary_keys = set(summary_df["品名"])
    unmatched_keys = sorted(list(forecast_keys - summary_keys))

    return result, unmatched_keys

def merge_forecast_header(sheet):
    """
    自动检测以“月预测”结尾的列（如“6月预测”、“7月预测”），
    在第一行合并这些列的单元格并写入“预测”，设置居中。
    """
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # 找到所有“月预测”结尾的列索引
    forecast_cols = [
        idx for idx, col in enumerate(header_row, start=1)
        if isinstance(col, str) and col.endswith("月预测")
    ]

    if not forecast_cols:
        return  # 没有预测列，不处理

    start_col = min(forecast_cols)
    end_col = max(forecast_cols)

    # 合并单元格
    merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
    sheet.merge_cells(merge_range)

    # 设置内容与样式
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "预测"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
def merge_finished_inventory_with_warehouse_types(summary_df: pd.DataFrame, finished_inventory_df: pd.DataFrame, mapping_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    1. 提取成品库存的“HOLD仓”、“成品仓”、“半成品仓”库存数量，根据“品名”合并进主计划；
    2. 对于 mapping_df 中“新品名”对应的“半成品”，如果“半成品”在库存表中且仓库为“半成品仓”，则其数量加到新品名的“半成品仓”。

    返回：
    - 合并后的主计划 DataFrame
    - 未匹配品名列表（仅成品库存中存在但主计划中不存在）
    """    
    warehouse_cols = ["HOLD仓", "成品仓", "半成品仓"]

    # 初始化列
    for col in warehouse_cols:
        if col not in summary_df.columns:
            summary_df[col] = 0

    # === 清洗数据 ===
    finished_df = finished_inventory_df.copy()
    finished_df["品名"] = finished_df["品名"].astype(str).str.strip()
    finished_df["仓库名称"] = finished_df["仓库名称"].astype(str).str.strip()
    finished_df["数量"] = pd.to_numeric(finished_df["数量"], errors="coerce").fillna(0)

    summary_df["品名"] = summary_df["品名"].astype(str).str.strip()

    # === 主逻辑：直接匹配仓库类型 ===
    grouped = finished_df.groupby(["品名", "仓库名称"], as_index=False)["数量"].sum()

    for _, row in grouped.iterrows():
        pname = row["品名"]
        warehouse = row["仓库名称"]
        qty = row["数量"]

        if warehouse in warehouse_cols and pname in summary_df["品名"].values:
            summary_df.loc[summary_df["品名"] == pname, warehouse] += qty
            
    # === 处理半成品映射逻辑 ===
    mapping_df = mapping_df.copy()
    mapping_df["半成品"] = mapping_df["半成品"].astype(str).str.strip()
    mapping_df["新品名"] = mapping_df["新品名"].astype(str).str.strip()

    # 只保留半成品列不为空的映射
    valid_mappings = mapping_df[mapping_df["半成品"] != ""]

    for _, row in valid_mappings.iterrows():
        old_name = row["半成品"]
        new_name = row["新品名"]

        if old_name and new_name:
            # 查找 old_name 的“半成品仓”库存
            match = grouped[
                (grouped["品名"] == old_name) & (grouped["仓库名称"] == "半成品仓")
            ]

            if not match.empty:
                qty = match["数量"].sum()
                if new_name in summary_df["品名"].values:
                    summary_df.loc[summary_df["品名"] == new_name, "半成品仓"] += qty

    # === 返回未匹配品名（成品库存中存在但主计划中不存在）===
    unmatched = sorted(list(set(finished_df["品名"]) - set(summary_df["品名"])))

    return summary_df, unmatched

def merge_inventory_header(sheet):
    """
    合并“HOLD仓”、“成品仓”、“半成品仓”标题，写入“库存”，居中。
    """
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    inventory_cols = [
        idx for idx, col in enumerate(header_row, start=1)
        if col in ["HOLD仓", "成品仓", "半成品仓"]
    ]
    if not inventory_cols:
        return

    start_col = min(inventory_cols)
    end_col = max(inventory_cols)
    sheet.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "成品库存"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_product_in_progress(summary_df: pd.DataFrame,
                               product_in_progress_df: pd.DataFrame,
                               mapping_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    将成品在制表中数据按“品名”合并进主计划表：
    - 半成品通过 mapping_df 中“半成品”映射到“新品名”，填入“半成品在制”列；
    - 其他数据直接匹配“产品品名” → “成品在制”列；
    返回合并后的表格与未匹配的品名列表。
    """
    summary_df = summary_df.copy()
    summary_df["成品在制"] = 0
    summary_df["半成品在制"] = 0

    # 数值列：只处理数值型的未交列
    numeric_cols = product_in_progress_df.select_dtypes(include='number').columns.tolist()
    if "未交" not in product_in_progress_df.columns:
        raise ValueError("❌ '成品在制'文件中未找到 '未交' 列")
    
    product_in_progress_df["产品品名"] = product_in_progress_df["产品品名"].astype(str).str.strip()
    summary_df["品名"] = summary_df["品名"].astype(str).str.strip()
    mapping_df["半成品"] = mapping_df["半成品"].astype(str).str.strip()
    mapping_df["新品名"] = mapping_df["新品名"].astype(str).str.strip()

    used_keys = set()
    unmatched_keys = set()

    # === 处理半成品在制 ===
    semi_rows = mapping_df[mapping_df["半成品"] != ""]
    matched_half = product_in_progress_df[
        product_in_progress_df["产品品名"].isin(semi_rows["半成品"])
    ]

    # 聚合半成品 → 新品名
    for _, row in semi_rows.iterrows():
        semi = row["半成品"]
        new = row["新品名"]
        value = matched_half.loc[
            matched_half["产品品名"] == semi, "未交"
        ].sum()

        if new in summary_df["品名"].values:
            summary_df.loc[summary_df["品名"] == new, "半成品在制"] += value
            used_keys.add(new)
        else:
            unmatched_keys.add(new)

    # === 删除已处理的半成品行 ===
    remaining = product_in_progress_df[
        ~product_in_progress_df["产品品名"].isin(semi_rows["半成品"])
    ]

    # === 处理成品在制 ===
    for _, row in remaining.iterrows():
        pname = row["产品品名"]
        qty = row["未交"]
        if pname in summary_df["品名"].values:
            summary_df.loc[summary_df["品名"] == pname, "成品在制"] += qty
            used_keys.add(pname)
        else:
            unmatched_keys.add(pname)

    return summary_df, sorted(list(unmatched_keys - used_keys))

def merge_product_in_progress_header(sheet):
    """
    合并“成品在制”“半成品在制”列，在第一行写入“成品在制”，居中。
    """
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cols = [
        idx for idx, val in enumerate(header_row, start=1)
        if val in ["成品在制", "半成品在制"]
    ]

    if not cols:
        return

    start_col = min(cols)
    end_col = max(cols)

    sheet.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "成品在制"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_order_delivery_amount_columns(main_plan_df: pd.DataFrame,
                               df_price: pd.DataFrame, start_date: datetime = None) -> tuple[pd.DataFrame, list]:
    """
    添加两列：
    - 匹配到当月订单可发货金额
    - 匹配到所有订单可发货金额
    逻辑：
    - 当月发货金额：min(当月未交, 成品仓) × 单价
    - 所有发货金额：min(总未交订单, 成品仓) × 单价
    """
    if df_price is None or df_price.empty:
        st.warning("⚠️ 未交订单为空，无法提取单价")
        main_plan_df["匹配到当月订单可发货金额"] = 0
        main_plan_df["匹配到所有订单可发货金额"] = 0
        main_plan_df["订单外可发货金额"] = 0
        return main_plan_df

    # 提取品名到单价（注意清洗）
    name_col = "品名"
    price_col = "单价-原币"
    
    # ✅ 检查是否存在“单价-原币”列
    if price_col not in df_price.columns:
        st.error(f"❌ 当前上传的未交订单表中缺少 `{price_col}` 列，请检查列名是否正确。")
        st.stop()  # 中断后续运行，避免崩溃
    
    # ✅ 正常处理
    df_price[name_col] = df_price[name_col].astype(str).str.strip()
    price_map = (
        df_price
        .dropna(subset=[price_col])
        .groupby(name_col)[price_col]
        .mean()
        .to_dict()
    )

    # 获取字段
    today = pd.Timestamp(start_date.replace(day=1)) if start_date else pd.Timestamp(datetime.today().replace(day=1))
    current_month = today.strftime("%Y-%m")
    current_month_col = f"未交订单 {current_month}"
    total_unfulfilled_col = "总未交订单"
    inventory_col = "成品仓"

    # 若字段不存在，添加 0 值
    for col in [current_month_col, total_unfulfilled_col, inventory_col]:
        if col not in main_plan_df.columns:
            main_plan_df[col] = 0

    # 匹配金额列初始化
    current_delivery = []
    total_delivery = []
    additional_delivery = []

    for _, row in main_plan_df.iterrows():
        name = str(row["品名"]).strip()
        price = price_map.get(name, 0)

        unfulfilled_curr = row.get(current_month_col, 0)
        unfulfilled_total = row.get(total_unfulfilled_col, 0)
        inventory = row.get(inventory_col, 0)

        amt_curr = min(unfulfilled_curr, inventory) * price
        amt_total = min(unfulfilled_total, inventory) * price

        if inventory > unfulfilled_total:
            amt_additional = (inventory - unfulfilled_total) * price
        else:
            amt_additional = 0

        current_delivery.append(round(amt_curr, 2))
        total_delivery.append(round(amt_total, 2))
        additional_delivery.append(round(amt_additional, 2))

    main_plan_df["匹配到当月订单可发货金额"] = current_delivery
    main_plan_df["匹配到所有订单可发货金额"] = total_delivery
    main_plan_df["订单外可发货金额"] = additional_delivery
    
    return main_plan_df

def merge_order_delivery_amount(sheet):
    """
    合并“匹配到当月订单可发货金额”“匹配到所有订单可发货金额”列，在第一行写入“发货金额”，居中。
    """
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cols = [
        idx for idx, val in enumerate(header_row, start=1)
        if val in ["匹配到当月订单可发货金额", "匹配到所有订单可发货金额", "订单外可发货金额"]
    ]

    if not cols:
        return

    start_col = min(cols)
    end_col = max(cols)

    sheet.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "发货金额"
    cell.alignment = Alignment(horizontal="center", vertical="center")

def append_forecast_accuracy_column(main_plan_df: pd.DataFrame, start_date: datetime = None) -> pd.DataFrame:
    """
    在“半成品在制”后面插入一列：当月预测准确率(订单/预测)
    逻辑：
    - 若 (未交 + 销售) == 0 且 预测 > 0，则准确率 = -9999
    - 若 (未交 + 销售) > 0 且 预测 == 0，则准确率 = 9999
    - 若 (未交 + 销售) > 0 且 预测 > 0，则准确率 = (未交 + 销售) / 预测 × 100%
    """
    today = pd.Timestamp(start_date.replace(day=1)) if start_date else pd.Timestamp(datetime.today().replace(day=1))
    current_year = today.strftime("%Y-%m")
    current_month = str(today.month)  # 输出 "6"


    forecast_col = f"{current_month}月预测"
    unfulfilled_col = f"未交订单 {current_year}"
    fulfilled_col = f"{current_month}月销售数量"
    accuracy_col = "当月预测准确率(订单/预测)"
    
    
    # 填充缺失值为0
    forecast = main_plan_df[forecast_col].fillna(0)
    unfulfilled = main_plan_df[unfulfilled_col].fillna(0)
    fulfilled = main_plan_df[fulfilled_col].fillna(0)
    
    total_order = unfulfilled + fulfilled

    # 应用逻辑（向量化计算）
    accuracy = pd.Series(index=main_plan_df.index, dtype=object)
    mask1 = (total_order == 0) & (forecast > 0)
    mask2 = (total_order > 0) & (forecast == 0)
    mask3 = (total_order > 0) & (forecast > 0)

    accuracy[mask1] = -9999
    accuracy[mask2] = 9999
    accuracy[mask3] = ((total_order[mask3] / forecast[mask3]) * 100).round(1).astype(str) + "%"
    
    # 插入到“半成品在制”后面
    col_names = list(main_plan_df.columns)
    try:
        insert_pos = col_names.index("半成品在制") + 1
        main_plan_df = pd.concat([
            main_plan_df.iloc[:, :insert_pos],
            pd.DataFrame({accuracy_col: accuracy}),
            main_plan_df.iloc[:, insert_pos:]
        ], axis=1)
    except ValueError:
        # 若找不到就添加到最后
        main_plan_df[accuracy_col] = accuracy
        
    return main_plan_df

def merge_forecast_accuracy(sheet):
    """
    合并“匹配到当月订单可发货金额”“匹配到所有订单可发货金额”列，在第一行写入“发货金额”，居中。
    """
    header_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    cols = [
        idx for idx, val in enumerate(header_row, start=1)
        if val in ["当月预测准确率(订单/预测)"]
    ]

    if not cols:
        return

    start_col = min(cols)
    end_col = max(cols)

    sheet.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")
    cell = sheet.cell(row=1, column=start_col)
    cell.value = "预测准确率"
    cell.alignment = Alignment(horizontal="center", vertical="center")
