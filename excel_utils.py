import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import numbers
from openpyxl.styles import Font

def adjust_column_width(ws, max_width=70):
    """
    自动调整工作表中每列的宽度，适配内容长度。
    忽略第一行（通常用于填充非数据内容），以第二行起为基准。
    """
    for col_cells in ws.iter_cols(min_row=2):  # 从第二行开始
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 8, max_width)

def highlight_replaced_names_in_main_sheet(ws, replaced_names: list[str], name_col_header: str = "品名", header_row_idx: int = 2):
    """
    只标红主计划 sheet 中所有品名在 replaced_names 中的整行的前三列。

    参数：
        ws: openpyxl 的 worksheet 对象（主计划）
        replaced_names: 替换过的新名字列表
        name_col_header: 表头中品名字段名称，默认是“品名”
        header_row_idx: 表头所在的行号（默认第 2 行）
    """
    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

    # 获取表头行
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row_idx]]

    if name_col_header not in header_row:
        raise ValueError(f"❌ 未找到“{name_col_header}”列，无法标红替换新品名。")

    name_col_idx = header_row.index(name_col_header) + 1  # openpyxl 列从 1 开始

    # 遍历数据行（从 header 下一行开始）
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        cell_value = str(row[name_col_idx - 1].value).strip()
        if cell_value in replaced_names:
            for cell in row[:3]:  # 只标红前三列 A, B, C
                cell.fill = red_fill

def reorder_main_plan_by_unfulfilled_sheet(main_plan_df: pd.DataFrame, unfulfilled_df: pd.DataFrame, name_col: str = "品名") -> pd.DataFrame:
    """
    根据“未交订单汇总”中的品名顺序对主计划进行排序，优先将这些品名排在前面。

    参数：
        main_plan_df: 主计划 DataFrame
        unfulfilled_df: 未交订单汇总 DataFrame
        name_col: 品名列名，默认“品名”

    返回：
        排序后的主计划 DataFrame
    """
    if name_col not in main_plan_df.columns or name_col not in unfulfilled_df.columns:
        raise ValueError(f"❌ 主计划或未交订单中缺少列：{name_col}")

    # 获取未交订单中品名的顺序列表
    priority_names = unfulfilled_df[name_col].dropna().astype(str).str.strip().unique().tolist()

    # 添加排序键
    main_plan_df["_排序键"] = main_plan_df[name_col].astype(str).str.strip().apply(
        lambda x: priority_names.index(x) if x in priority_names else len(priority_names)
    )

    # 按排序键排序
    main_plan_df = main_plan_df.sort_values(by="_排序键").drop(columns="_排序键").reset_index(drop=True)

    return main_plan_df

def format_currency_columns_rmb(ws: Worksheet):
    """
    将所有标题中包含“金额”的列设置为人民币格式（¥#,##0.00），自动转换字符串为数字
    """
    header_row = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(1, max_col + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if isinstance(header, str) and "金额" in header:
            for row_idx in range(header_row + 1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                # 判断是否为数字或可转为数字的字符串
                if isinstance(val, (int, float)):
                    cell.number_format = u'¥#,##0.00'
                elif isinstance(val, str):
                    try:
                        num = float(val.replace(",", "").strip())
                        cell.value = num
                        cell.number_format = u'¥#,##0.00'
                    except:
                        pass  # 非法字符串忽略

def format_thousands_separator(ws: Worksheet):
    """
    将 F 列（第6列）之后的所有列，若为数值或可转数字的字符串，设置为千位分隔格式：#,##0.00。
    """
    header_row = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(6, max_col + 1):  # 从F列开始
        for row_idx in range(header_row + 1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value

            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
            elif isinstance(val, str):
                try:
                    num = float(val.replace(",", "").strip())
                    cell.value = num
                    cell.number_format = '#,##0'
                except:
                    continue

def add_sheet_hyperlinks(ws: Worksheet, sheet_names: list):
    """
    给 ws 的第二列添加超链接，跳转到名称相同的工作表。
    要求第 1 行是表头，从第 2 行开始为内容。

    参数：
        ws: openpyxl 的工作表对象（即“图”）
        sheet_names: 所有已存在的 sheet 名称列表
    """
    hyperlink_font = Font(color="0000FF", underline="single")  # 蓝色 + 下划线

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        target_sheet = cell.value
        if target_sheet and target_sheet in sheet_names:
            # 添加内部超链接
            cell.value = f'=HYPERLINK("#\'{target_sheet}\'!A1", "{target_sheet}")'
            cell.font = hyperlink_font  # 设置样式
