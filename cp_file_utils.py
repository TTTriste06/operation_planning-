import re
import pandas as pd
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Alignment

def merge_cp_files_by_keyword(cp_dataframes: dict) -> dict:
    grouped = defaultdict(list)

    # 将 DB, DB2, DB3... 聚合成同一组
    for key, df in cp_dataframes.items():
        for kw in ["华虹", "先进", "DB", "上华1厂", "上华2厂", "上华5厂"]:
            if key.startswith(kw):
                grouped[kw].append(df)
                break

    # 合并
    merged_cp_dataframes = {}
    for kw, df_list in grouped.items():
        # 排除空 DataFrame
        df_list = [df for df in df_list if df is not None and not df.empty]
        if df_list:
            merged_cp_dataframes[kw] = pd.concat(df_list, ignore_index=True)
        else:
            merged_cp_dataframes[kw] = pd.DataFrame()

    return merged_cp_dataframes

def extract_month_week(s):
    match = re.match(r"(\d{4})-(\d{2}) WK(\d)", s)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        week = int(match.group(3))
        return (year, month, week)
    return (9999, 99, 99)  # 排最后

def generate_fab_summary(cp_dataframes: dict) -> pd.DataFrame:
    import pandas as pd

    fab_rules = {
        "上华1厂": {"key": "上华1厂", "fab": "CSMC-1", "part": "CUST_PARTNAME", "qty": "CURRENT_QTY", "date": "FORECAST_FAB_OUT_DATE"},
        "上华2厂": {"key": "上华2厂", "fab": "CSMC-2", "part": "CUST_PARTNAME", "qty": "CURRENT_QTY", "date": "FORECAST_FAB_OUT_DATE"},
        "上华5厂": {"key": "上华5厂", "fab": "CSMC-5", "part": "CUST_PARTNAME", "qty": "CURRENT_QTY", "date": "FORECAST_FAB_OUT_DATE"},
        "DB":     {"key": "DB", "fab": "DB", "part": "Customer Device", "qty": "Cur Wfs", "date": "Confirmed Date"},
        "华虹":    {"key": "华虹", "fab": "HHG", "part": "客户品名", "qty": "当前数量", "date": "最终确定交货日期"},
        "先进积塔": {"key": "先进", "fab": "ASMC-GTA", "part": "Device ID", "qty": "End Qty", "date": "Estimate Out Date"},
    }

    def get_week_label(dt: pd.Timestamp) -> str:
        if pd.isnull(dt):
            return None
        year_month = dt.strftime("%Y-%m")
        day = dt.day
        if 1 <= day <= 7:
            return f"{year_month} WK1(1–7)"
        elif 8 <= day <= 15:
            return f"{year_month} WK2(8–15)"
        elif 16 <= day <= 22:
            return f"{year_month} WK3(16–22)"
        else:
            return f"{year_month} WK4(23–end)"

    all_rows = []

    for label, rule in fab_rules.items():
        fab_key = rule["key"]
        fab_value = rule["fab"]
        part_col = rule["part"]
        qty_col = rule["qty"]
        date_col = rule["date"]

        # 匹配所有相关表
        for sheet_name, df in cp_dataframes.items():
            if fab_key in sheet_name:
                if not all(col in df.columns for col in [part_col, qty_col, date_col]):
                    continue  # 跳过缺列的表

                df_temp = df[[part_col, qty_col, date_col]].copy()
                df_temp.columns = ["晶圆型号", "数量", "出货日期"]
                df_temp["出货日期"] = pd.to_datetime(df_temp["出货日期"], errors="coerce")
                df_temp["FAB"] = fab_value
                df_temp["周"] = df_temp["出货日期"].apply(get_week_label)
                all_rows.append(df_temp)

    df_all = pd.concat(all_rows, ignore_index=True)
    df_all = df_all.dropna(subset=["周", "数量", "晶圆型号"])

    # 透视表：晶圆型号 + FAB 行，列为每周，值为数量之和
    result = pd.pivot_table(
        df_all,
        index=["晶圆型号", "FAB"],
        columns="周",
        values="数量",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # 列顺序排序
    week_cols = sorted(
        [col for col in result.columns if isinstance(col, str) and "WK" in col],
        key=extract_month_week
    )
    result = result[["晶圆型号", "FAB"] + week_cols]


    return result

def format_fab_summary_month_headers(ws):
    """
    格式化 FAB_WIP_汇总 表头：
    - 第 1 行按年份-月份（如 2025-07）合并相同月份列块
    - 第 2 行去除列名前缀（只保留 WK 信息）
    - 第 1/2 行按月统一填充颜色
    """
    fill_colors = [
        "FFF2CC",  # 浅黄
        "D9EAD3",  # 浅绿
        "CFE2F3",  # 浅蓝
        "F4CCCC",  # 浅红
        "EAD1DC",  # 浅紫
        "D9D2E9",  # 浅灰紫
        "FCE5CD",  # 淡橘
        "D0E0E3"   # 灰蓝
    ]

    max_col = ws.max_column
    month_positions = {}  # {2025-07: (start_col, end_col)}
    month_order = []

    for col in range(3, max_col + 1):  # 从第3列开始是月份周数据
        cell = ws.cell(row=2, column=col)
        value = str(cell.value)

        # 提取年月信息：形如 "2025-07 WK2(8–15)"
        match = re.match(r"(\d{4}-\d{2}) WK\d", value)
        if match:
            year_month = match.group(1)
            week_info = value.split(" ", 1)[-1]  # 保留 "WK2(8–15)" 等

            # 更新 cell 显示值：去掉年份月份前缀
            cell.value = week_info

            if year_month not in month_positions:
                month_positions[year_month] = [col, col]
                month_order.append(year_month)
            else:
                month_positions[year_month][1] = col

    # 插入第一行标题并合并、着色
    for idx, month in enumerate(month_order):
        start_col, end_col = month_positions[month]
        color = fill_colors[idx % len(fill_colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # 第1行合并并写入月份
        cell = ws.cell(row=1, column=start_col)
        cell.value = month
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if start_col != end_col:
            ws.merge_cells(start_row=1, end_row=1, start_column=start_col, end_column=end_col)

        # 为该月范围内的两行着色
        for col in range(start_col, end_col + 1):
            ws.cell(row=1, column=col).fill = fill
            ws.cell(row=2, column=col).fill = fill

    # 清空前两列第1行
    ws.cell(row=1, column=1).value = ""
    ws.cell(row=1, column=2).value = ""
